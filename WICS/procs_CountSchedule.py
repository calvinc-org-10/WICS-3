import datetime
import os, uuid
import subprocess, signal
import re as regex
import json
from django.conf import settings as django_settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Value, Subquery, OuterRef
from django.db.models.query import QuerySet
from django.http import HttpResponse, HttpRequest
from django.shortcuts import render
from django.template.loader import render_to_string
from django.views.generic import ListView
from django_q.tasks import async_task
from barcode import Code128
from userprofiles.models import WICSuser
from cMenu.models import getcParm
from cMenu.utils import makebool, WrapInQuotes
from cMenu.utils import ExcelWorkbook_fileext
from cMenu.utils import isDate, calvindate
from cMenu.utils import user_db
from WICS.models import MaterialList, VIEW_materials, CountSchedule, VIEW_countschedule, VIEW_LastFoundAtList
from WICS.models_async_comm import async_comm, set_async_comm_state
from WICS.procs_SAP import fnSAPList
from typing import *
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH
# below: skip Sat, Sun using dateutil, dateutil.rrule, dateutil.rruleset
# implement skipping holidays
from cMenu.utils import calvindate
# from WICS.procs_misc import HolidayList


# ====================================================
# ====================================================
# ====================================================


def fnCountScheduleRecordExists(req:HttpRequest, CtDate:datetime, Matl:MaterialList|str|int) -> CountSchedule:
    """
    used to check if a CountSchedule record exists for the given CtDate and Material
    (only one such record is allowed).
    If the record exists, it is the return value, else a None CountSchedule rec
    """
    if isinstance(Matl,MaterialList):
        MatObj = Matl 
    elif isinstance(Matl,str):
        try:
            MatObj = MaterialList.objects.using(user_db(req)).get(Material=Matl)
        except:
            MatObj = MaterialList.objects.using(user_db(req)).none()
    elif isinstance(Matl,int):
        try:
            MatObj = MaterialList.objects.using(user_db(req)).get(pk=Matl)
        except:
            MatObj = MaterialList.objects.using(user_db(req)).none()
    else:
        MatObj = MaterialList.objects.using(user_db(req)).none()
        
    try:
        rec = CountSchedule.objects.using(user_db(req)).get(CountDate=CtDate, Material=MatObj)
    except:
        rec = CountSchedule.objects.using(user_db(req)).none()

    return rec

#####################################################################
#####################################################################
#####################################################################

@login_required
def fnUploadCountSchedSprsht(req):

    SprshtDateEpoch = WINDOWS_EPOCH

    def cleanupfld(fld, val):
        """
        fld is the name of the field in the ActualCount or MaterialList table
        val is the value to be cleaned for insertion into the fld
        Returns  {'usefld':usefld, 'cleanval': cleanval}
            usefld is a boolean indicating that val could/not be cleaned to the correct type
            cleanval is val in the correct type (if usefld==True)
        """
        cleanval = None

        if   fld == 'CountDate': 
            if isinstance(val,(calvindate, datetime.date, datetime.datetime)):
                usefld = True
                cleanval = calvindate(val).as_datetime()
            elif isinstance(val,int):
                usefld = True
                cleanval = from_excel(val,SprshtDateEpoch)
            else:
                usefld = isDate(val) 
                if (usefld != False):
                    cleanval = calvindate(usefld).as_datetime()
                    usefld = True
        elif fld in \
            ['org_id', 
            ]:
            try:
                cleanval = int(val)
                usefld = True
            except:
                usefld = False
        elif fld in \
            ['Material', 
             'Counter', 
             'Notes', 
             'Priority'
             'ReasonScheduled',
            ]:
            usefld = (val is not None)
            if usefld: cleanval = str(val)
        else:
            usefld = True
            cleanval = val
        
        return {'usefld':usefld, 'cleanval': cleanval}
    #end def cleanupfld

    if req.method == 'POST':
        UplResults = []
        nRowsAdded = 0
        SprshtRowNum = 0

        # save the file so we can open it as an excel file
        SprshtFile = req.FILES['SchdFile']
        svdir = getcParm(req, 'SAP-FILELOC')
        fName = svdir+"tmpSchd"+str(uuid.uuid4())+ExcelWorkbook_fileext
        with open(fName, "wb") as destination:
            for chunk in SprshtFile.chunks():
                destination.write(chunk)

        wb = load_workbook(filename=fName, read_only=True)
        SprshtDateEpoch = wb.epoch
        if 'Schedule' in wb:
            ws = wb['Schedule']
        else:
            UplResults.append({'error':'This workbook does not contain a sheet named Schedule in the correct format'})
            ws = None

        if ws:
            SprshtcolmnNames = ws[1]
            SprshtREQUIREDFLDS = ['org_id','Material','CountDate']     
            SprshtcolmnMap = {}
            Sprsht_SSName_TableName_map = {
                    'org_id': 'org_id',         # org_id and Material will be converted to a Material_id
                    'Material': 'Material',     # org_id and Material will be converted to a Material_id
                    
                    'CountDate': 'CountDate',
                    'Counter': 'Counter',
                    'Priority': 'Priority',
                    'ReasonScheduled': 'ReasonScheduled',
                    'Notes': 'Notes',
                    }
            for col in SprshtcolmnNames:
                if col.value in Sprsht_SSName_TableName_map:
                    SprshtcolmnMap[Sprsht_SSName_TableName_map[col.value]] = col.column - 1
            
            HeaderGood = True
            for reqFld in SprshtREQUIREDFLDS:
                HeaderGood = HeaderGood and (reqFld in SprshtcolmnMap)
            if not HeaderGood:
                UplResults.append({'error':'The Schedule worksheet in this workbook has bad header row'})
                ws = None

        if ws:
            SprshtRowNum=1
            MAX_COUNT_ROWS = 5000   # remove if this proc is made asynchronous
            for row in ws.iter_rows(min_row=SprshtRowNum+1, max_row=MAX_COUNT_ROWS, values_only=True):
                SprshtRowNum += 1

                # if no org given, check that Material unique.
                if 'org_id' not in SprshtcolmnMap:
                    spshtorg = None
                else:
                    spshtorg = cleanupfld('org_id', row[SprshtcolmnMap['org_id']])['cleanval']
                matlnum = cleanupfld('Material', row[SprshtcolmnMap['Material']])['cleanval']
                matlorglist = MaterialList.objects.using(user_db(req)).filter(Material=matlnum).values_list('org_id', flat=True)
                MatlKount = len(matlorglist)
                MatObj = None
                err_already_handled = False
                if MatlKount == 1:
                    MatObj = MaterialList.objects.using(user_db(req)).get(Material=matlnum)
                    spshtorg = MatObj.org_id
                if MatlKount > 1:
                    if spshtorg is None:
                        UplResults.append({
                            'error': f"{matlnum} in multiple org_id's {tuple(matlorglist)}, but no org_id given", 
                            'rowNum': SprshtRowNum,
                            })
                        err_already_handled = True
                    elif spshtorg in matlorglist:
                        MatObj = MaterialList.objects.using(user_db(req)).get(org_id=spshtorg, Material=matlnum)
                    else:
                        UplResults.append({
                            'error': f"{matlnum} in in multiple org_id's {tuple(matlorglist)}, but org_id given ({spshtorg}) is not one of them", 
                            'rowNum': SprshtRowNum,
                            })
                        err_already_handled = True
                    #endif spshtorg is None
                #endif MatKount > 1

                if matlnum and not MatObj:
                    if not err_already_handled:
                        UplResults.append({'error':'either ' + matlnum + ' does not exist in MaterialList or incorrect org_id (' + str(spshtorg) + ') given', 'rowNum':SprshtRowNum})
                elif matlnum and MatObj:
                    requiredFields = {reqFld: False for reqFld in SprshtREQUIREDFLDS}
                    # org_id and Material must be present, else there'd be no MatObj
                    requiredFields['org_id'] = True
                    requiredFields['Material'] = True

                    SRec = CountSchedule()
                    for fldName, colNum in SprshtcolmnMap.items():
                        # check/correct problematic data types
                        usefld, V = cleanupfld(fldName, row[colNum]).values()
                        if (V is not None):
                            if usefld:
                                if   fldName == 'CountDate': 
                                    setattr(SRec, fldName, V) 
                                    requiredFields['CountDate'] = True
                                elif fldName == 'Material': 
                                    setattr(SRec, fldName, MatObj)
                                    requiredFields['Material'] = True
                                elif fldName == 'Counter': 
                                    setattr(SRec, fldName, V)
                                    requiredFields['Counter'] = True
                                else:
                                    if hasattr(SRec, fldName): setattr(SRec, fldName, V)
                                #endif fldName == ...
                            else:
                                UplResults.append({'error':str(V)+' is invalid for '+fldName, 'rowNum':SprshtRowNum})
                            #endif usefld
                        #endif (V is not None)
                    #endfor fldName, colNum

                    # are all required fields present?
                    AllRequiredPresent = True
                    for keyname, Prsnt in requiredFields.items():
                        AllRequiredPresent = AllRequiredPresent and Prsnt
                        if not Prsnt:
                            UplResults.append({'error':keyname+' missing', 'rowNum':SprshtRowNum})

                    if AllRequiredPresent:
                        if fnCountScheduleRecordExists(req, SRec.CountDate, MatObj):
                            UplResults.append({'error': f'Count already scheduled for {MatObj}  on {SRec.CountDate:%Y-%m-%d}', 'rowNum':SprshtRowNum})
                        else:
                            SRec.save(using=user_db(req))
                            qs = type(SRec).objects.using(user_db(req)).filter(pk=SRec.pk).values().first()
                            res = {'error': False, 'rowNum':SprshtRowNum, 'MaterialNum': str(MatObj) }
                            res.update(qs)      # tack the new record (along with its new pk) onto res
                            # res.update(SRec)      #QUESTION:  can I do this directly with SRec?? - NO. This line does not work. SRec is not iterable
                            UplResults.append(res)
                            nRowsAdded += 1
                #endif matlnum and MatObj/not MatObj

            if SprshtRowNum >= MAX_COUNT_ROWS:
                UplResults.insert(0,{'error':f'Data in spreadsheet rows {MAX_COUNT_ROWS+1} and beyond are being ignored.'})
        #endif ws

        # close and kill temp files
        wb.close()
        os.remove(fName)

        cntext = {
            'UplResults':UplResults, 
            'nRowsRead':SprshtRowNum, 
            'nRowsAdded':nRowsAdded,
                }
        templt = 'frm_uploadSchedCount_Success.html'
    else:
        cntext = {
                }
        templt = 'frm_UploadSchedCountSprdsht.html'
    #endif

    return render(req, templt, cntext)


#####################################################################
#####################################################################
#####################################################################

class CountScheduleListForm(LoginRequiredMixin, ListView):
    ordering = ['-CountDate', 'Material']
    context_object_name = 'CtSchdList'
    template_name = 'frm_CountScheduleList.html'
    
    def setup(self, req: HttpRequest, *args: Any, **kwargs: Any) -> None:
        self._user = req.user
        return super().setup(req, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Any]:
        return VIEW_countschedule(self._user).order_by(*self.ordering)

    # def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return super().get(request, *args, **kwargs)

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        return HttpResponse('Stop rushing me!!')

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        cntext = super().get_context_data(**kwargs)

        return cntext


#####################################################################
#####################################################################
#####################################################################

class CountWorksheetReport(LoginRequiredMixin, ListView):
    ordering = ['Counter', 'Material__Material']
    context_object_name = 'CtSchd'
    template_name = 'rpt_CountWksht_main.html'
    
    # def setup(self, userid, reqid, *args: Any, **kwargs: Any) -> None:
    def setup(self, db_to_use: str, reqid:int, *args: Any, **kwargs: Any) -> None:
        self._db = db_to_use
        self._reqid = reqid
        if 'CountDate' in kwargs: self.CountDate = kwargs['CountDate']
        else: self.CountDate = calvindate().today().as_datetime()
        if isinstance(self.CountDate,str): self.CountDate = calvindate(self.CountDate).as_datetime()
        self.SAPDate = fnSAPList(db_to_use, self.CountDate)['SAPDate']
        # return super().setup(req, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Any]:
        SAP_SOH = fnSAPList(self._db, self.CountDate)
        self.SAPDate = SAP_SOH['SAPDate']
        qs = CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate).order_by(*self.ordering).select_related('Material','Material__PartType')
        qs = qs.annotate(LastCountDate=Value(0), LastFoundAt=Value(''), SAPQty=Value(0), MaterialBarCode=Value(''), Material_org=Value(''))
        prevCtr = None
        for rec in qs:
            strMatlNum = str(rec.Material)
            rec.Material_org = strMatlNum
            set_async_comm_state(
                self._db,
                self._reqid,
                statecode = 'proc-Material',
                statetext = f'Preparing Count Worksheet for Material {strMatlNum}',
                )

            rec.prevCounter = prevCtr
            prevCtr = rec.Counter

            bcstr = Code128(str(strMatlNum)).render(writer_options={'module_height':7.0,'module_width':0.35,'quiet_zone':0.1,'write_text':True,'text_distance':3.5})
            bcstr = str(bcstr).replace("b'","").replace('\\r','').replace('\\n','').replace("'","")
            rec.MaterialBarCode = bcstr
            rec.LastFoundAt = VIEW_materials.objects.using(self._db).get(pk=rec.Material_id).LastFoundAt
            rec.LastCountDate = VIEW_materials.objects.using(self._db).get(pk=rec.Material_id).LastCountDate
            for SAProw in SAP_SOH['SAPTable'].filter(Material=rec.Material):
                rec.SAPQty += SAProw.Amount*SAProw.mult

        return qs

    # def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return super().get(request, *args, **kwargs)

    # there is no POST processing; it's a r/o report
    # def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return HttpResponse('Stop rushing me!!')

    def render_to_response(self) -> HttpResponse:
        context = {
            self.context_object_name: self.get_queryset()
        }

        # collect the list of Counters so that tabs can be built in the html
        set_async_comm_state(
            self._db,
            self._reqid,
            statecode = 'proc-CounterList',
            statetext = f'Collecting List of Counters',
            )
        CounterList = CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate).order_by('Counter').values('Counter').distinct()

        # form list of last found locations
        set_async_comm_state(
            self._db,
            self._reqid,
            statecode = 'proc-Locations',
            statetext = f'Collecting List of Locations',
            )
        MList = MaterialList.objects.using(self._db).filter(
                    pk__in=Subquery(CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate).values('Material'))
                )
        LocationList = VIEW_LastFoundAtList(self._db, MList).annotate(Counter=Subquery(CountSchedule.objects.filter(CountDate=self.CountDate,Material=OuterRef('Material')).values('Counter')[:1])).order_by('FoundAt') if MList.exists() else []

        context.update({
                'LocationList': LocationList,
                'CounterList': CounterList,
                'CountDate': self.CountDate,     # .as_datetime(), -- this was done in setup()
                'SAP_Updated_at': self.SAPDate,
                })

        return render_to_string(self.template_name,context)

##### the suite of procs to support viewCountWorksheetReport

def proc_CountWorksheet_00InitUMLasync_comm(req: HttpRequest, reqid: int) -> None:
    acomm = set_async_comm_state(
        user_db(req),
        reqid,
        statecode = 'initializing',
        statetext = 'Initializing ...',
        new_async=True
        )

def proc_CountWorksheet_01PrepWorksheet(db_to_use, reqid, CountDate):
    CountWorksheet_instance = CountWorksheetReport()
    CountWorksheet_instance.setup(db_to_use, reqid, CountDate=CountDate)
    resp = CountWorksheet_instance.render_to_response()

    set_async_comm_state(
        db_to_use,
        reqid,
        statecode = 'writing-tmpfile',
        statetext = f'Formatting Worksheet',
        )
    # write rendered_worksheet to temp file for pickup when client requests
    svdir = django_settings.STATIC_ROOT
    if svdir is None:
        svdir = django_settings.STATICFILES_DIRS[0]
    fName_base = '/tmpdl/'+'CWksht' + f'{reqid}'
    fName = svdir + fName_base
    if svdir is not None and fName_base is not None:
        with open(fName, "w") as destination:
            destination.write(str(resp))

    set_async_comm_state(
        db_to_use,
        reqid,
        statecode = 'done',
        statetext = f'Ready to present Worksheet',
        result=fName,
        )
    
def proc_CountWorksheet_99_Cleanup(db_to_use, reqid, tmpHTMLfil):
    # also kill reqid, acomm, qcluster process
    async_comm.objects.using(db_to_use).filter(pk=reqid).delete()

    # kill the django-q process
    try:
        os.kill(int(reqid), signal.SIGTERM)
    except AttributeError:
        pass
    try:
        os.kill(int(reqid), signal.SIGKILL)
    except AttributeError:
        pass

    # delete the temporary file  - NOPE = it's used after this proc ends
    # os.remove(tmpHTMLfil)


def viewCountWorksheetReport(req, CountDate=None):
    dbUsing = user_db(req)

    client_phase = req.POST['phase'] if 'phase' in req.POST else None
    reqid = req.COOKIES['reqid'] if 'reqid' in req.COOKIES else None

    # if CountDate is None: CountDate = calvindate().today().as_datetime()  # this is the old way - remove this from view params
    CountDate = req.POST['CountDate'] if 'CountDate' in req.POST else calvindate().today().as_datetime()

    if req.method == 'POST':
        retinfo = HttpResponse()
        if client_phase=='init-upl':
            # start django_q broker
            reqid = subprocess.Popen(
                ['python', f'{django_settings.BASE_DIR}/manage.py', 'qcluster']
            ).pid
            retinfo.set_cookie('reqid',str(reqid))

            proc_CountWorksheet_00InitUMLasync_comm(req, reqid)

            async_task(proc_CountWorksheet_01PrepWorksheet, dbUsing, reqid, CountDate)

            acomm = async_comm.objects.using(dbUsing).values().get(pk=reqid)    # something's very wrong if this doesn't exist
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='waiting':
            acomm = async_comm.objects.using(dbUsing).values().get(pk=reqid)    # something's very wrong if this doesn't exist
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='wantresults':
            # get results from temp file
            # write rendered_worksheet to temp file for pickup when client requests
            svdir = django_settings.STATIC_ROOT
            if svdir is None:
                svdir = django_settings.STATICFILES_DIRS[0]
            fName_base = '/tmpdl/'+'CWksht' + f'{reqid}'
            fName = svdir + fName_base
            if svdir is not None and fName_base is not None:
                with open(fName, "r") as readfile:
                    retHTML = readfile.read()

            # do final cleanup
            proc_CountWorksheet_99_Cleanup(dbUsing, reqid, fName)
            retinfo.delete_cookie('reqid')

            # return to requestor for presentation
            return retHTML
        elif client_phase=='cleanup-after-failure':
            pass
        else:
            return
        #endif client_phase
    else:
        # (hopefully,) this is the initial phase; all others will be part of a POST request

        SAPDate = fnSAPList(req, CountDate)['SAPDate']

        cntext = {
            'SAP_Updated_at': SAPDate,
            'CountDate': CountDate,
            }
        templt = 'rpt_CountWksht.html'
    #endif req.method = 'POST'

    return render(req, templt, cntext)
    pass

############################################################################
############################################################################
############################################################################

class CountWorksheetLocReport(LoginRequiredMixin, ListView):
    ordering = ['Counter', 'FoundAt', 'Material__Material']
    context_object_name = 'CtSchd'
    template_name = 'rpt_CountWkshtLoc_main.html'
    
    # def setup(self, userid, reqid, *args: Any, **kwargs: Any) -> None:
    def setup(self, db_to_use: str, reqid: int, *args: Any, **kwargs: Any) -> None:
        self._db  = db_to_use
        self._reqid = reqid
        if 'CountDate' in kwargs: self.CountDate = kwargs['CountDate']
        else: self.CountDate = calvindate().today().as_datetime()
        if isinstance(self.CountDate,str): self.CountDate = calvindate(self.CountDate).as_datetime()

        # return super().setup(req, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Any]:
        MList = MaterialList.objects.using(self._db).filter(
                    pk__in=Subquery(CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate).values('Material'))
                )
        qs = []
        if MList.exists():
            qs = VIEW_LastFoundAtList(self._db, MList).annotate(
                Counter=Subquery(CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate,Material=OuterRef('Material')).values('Counter')[:1]),
                MaterialBarCode=Value(''),
                ).order_by(*self.ordering) 

        # qs = CountSchedule.objects.filter(CountDate=self.CountDate).order_by(*self.ordering).select_related('Material','Material__PartType')
        # qs = qs.annotate(LastCountDate=Value(0), LastFoundAt=Value(''), SAPQty=Value(0), MaterialBarCode=Value(''), Material_org=Value(''))
        prevCtr = None
        prevLoc = None
        for rec in qs:
            strMatlNum = rec['Material_org']
            # rec.Material_org = strMatlNum
            set_async_comm_state(
                self._db,
                self._reqid,
                statecode = 'proc-Material',
                statetext = f'Preparing Count Worksheet for Material {strMatlNum} at {rec["FoundAt"]}',
                )

            rec['prevLoc'] = prevLoc
            rec['prevCounter'] = prevCtr
            prevCtr = rec['Counter']
            prevLoc = rec['FoundAt']

            bcstr = Code128(str(strMatlNum)).render(writer_options={'module_height':7.0,'module_width':0.35,'quiet_zone':0.1,'write_text':False})
            bcstr = str(bcstr).replace("b'","").replace('\\r','').replace('\\n','').replace("'","")
            rec['MaterialBarCode'] = bcstr

            bcstr = Code128(str(rec['FoundAt'])).render(writer_options={'module_height':7.0,'module_width':0.35,'quiet_zone':0.1,'write_text':False})
            bcstr = str(bcstr).replace("b'","").replace('\\r','').replace('\\n','').replace("'","")
            rec['LocationBarCode'] = bcstr

            rec['LastFoundAt'] = VIEW_materials.objects.get(pk=rec['Material']).LastFoundAt
            rec['LastCountDate'] = VIEW_materials.objects.get(pk=rec['Material']).LastCountDate
            
            # replace the Material number with the actual record
            rec['Material'] = MaterialList.objects.get(pk=rec['Material'])

        return qs

    # def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return super().get(request, *args, **kwargs)

    # there is no POST processing; it's a r/o report
    # def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return HttpResponse('Stop rushing me!!')

    def render_to_response(self) -> HttpResponse:
        Q = self.get_queryset()
        context = {
            self.context_object_name: Q
        }

        # collect the list of Counters so that tabs can be built in the html
        set_async_comm_state(
            self._db,
            self._reqid,
            statecode = 'proc-CounterList',
            statetext = f'Collecting List of Counters',
            )
        CounterList = CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate).order_by('Counter').values('Counter').distinct()

        # form Material summary list
        set_async_comm_state(
            self._db,
            self._reqid,
            statecode = 'proc-MatlSumm',
            statetext = f'Collecting Material Summary List',
            )
        MList = VIEW_materials.objects.using(self._db).filter(
                    pk__in=Subquery(CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate).values('Material'))
                ).annotate(
                    Counter=Subquery(CountSchedule.objects.using(self._db).filter(CountDate=self.CountDate,Material=OuterRef('Material')).values('Counter')[:1]),
                    MaterialBarCode=Value(''),
                ).order_by(
                    'Material', 'org'
                )
        for rec in MList:
            bcstr = Code128(str(rec.Material_org)).render(writer_options={'module_height':7.0,'module_width':0.35,'quiet_zone':0.1,'write_text':False})
            bcstr = str(bcstr).replace("b'","").replace('\\r','').replace('\\n','').replace("'","")
            rec.MaterialBarCode = bcstr

        context.update({
                'MaterialList': MList,
                'CounterList': CounterList,
                'CountDate': self.CountDate,     # .as_datetime(), -- this was done in setup()
                # 'SAP_Updated_at': self.SAPDate,
                })

        return render_to_string(self.template_name,context)

##### the suite of procs to support viewCountWorksheetLocReport

def proc_CountWorksheetLoc_00InitUMLasync_comm(db_to_use, reqid):
    acomm = set_async_comm_state(
        db_to_use,
        reqid,
        statecode = 'initializing',
        statetext = 'Initializing ...',
        new_async=True
        )

def proc_CountWorksheetLoc_01PrepWorksheet(db_to_use, reqid, CountDate):
    CountWorksheet_instance = CountWorksheetLocReport()
    CountWorksheet_instance.setup(db_to_use, reqid, CountDate=CountDate)
    resp = CountWorksheet_instance.render_to_response()

    set_async_comm_state(
        db_to_use,
        reqid,
        statecode = 'writing-tmpfile',
        statetext = f'Formatting Worksheet',
        )
    # write rendered_worksheet to temp file for pickup when client requests
    svdir = django_settings.STATIC_ROOT
    if svdir is None:
        svdir = django_settings.STATICFILES_DIRS[0]
    fName_base = '/tmpdl/'+'CWksht' + f'{reqid}'
    fName = svdir + fName_base
    if svdir is not None and fName_base is not None:
        with open(fName, "w") as destination:
            destination.write(str(resp))

    set_async_comm_state(
        db_to_use,
        reqid,
        statecode = 'done',
        statetext = f'Ready to present Worksheet',
        result=fName,
        )
    
def proc_CountWorksheetLoc_99_Cleanup(db_to_use, reqid, tmpHTMLfil):
    # also kill reqid, acomm, qcluster process
    async_comm.objects.using(user_db(db_to_use)).filter(pk=reqid).delete()

    # kill the django-q process
    try:
        os.kill(int(reqid), signal.SIGTERM)
    except AttributeError:
        pass
    try:
        os.kill(int(reqid), signal.SIGKILL)
    except AttributeError:
        pass

    # delete the temporary file  - NOPE = it's used after this proc ends
    # os.remove(tmpHTMLfil)


def viewCountWorksheetLocReport(req, CountDate=None):
    dbUsing = user_db(req)

    client_phase = req.POST['phase'] if 'phase' in req.POST else None
    reqid = req.COOKIES['reqid'] if 'reqid' in req.COOKIES else None
    # if CountDate is None: CountDate = calvindate().today().as_datetime()  # this is the old way - remove this from view params
    CountDate = req.POST['CountDate'] if 'CountDate' in req.POST else calvindate().today().as_datetime()

    if req.method == 'POST':
        retinfo = HttpResponse()
        if client_phase=='init-upl':
            # start django_q broker
            reqid = subprocess.Popen(
                ['python', f'{django_settings.BASE_DIR}/manage.py', 'qcluster']
            ).pid
            retinfo.set_cookie('reqid',str(reqid))

            proc_CountWorksheetLoc_00InitUMLasync_comm(dbUsing, reqid)

            async_task(proc_CountWorksheetLoc_01PrepWorksheet, dbUsing, reqid, CountDate)

            acomm = async_comm.objects.using(dbUsing).values().get(pk=reqid)    # something's very wrong if this doesn't exist
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='waiting':
            acomm = async_comm.objects.using(dbUsing).values().get(pk=reqid)    # something's very wrong if this doesn't exist
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='wantresults':
            # get results from temp file
            # write rendered_worksheet to temp file for pickup when client requests
            svdir = django_settings.STATIC_ROOT
            if svdir is None:
                svdir = django_settings.STATICFILES_DIRS[0]
            fName_base = '/tmpdl/'+'CWksht' + f'{reqid}'
            fName = svdir + fName_base
            if svdir is not None and fName_base is not None:
                with open(fName, "r") as readfile:
                    retHTML = readfile.read()

            # do final cleanup
            proc_CountWorksheetLoc_99_Cleanup(dbUsing, reqid, fName)
            retinfo.delete_cookie('reqid')

            # return to requestor for presentation
            return retHTML
        elif client_phase=='cleanup-after-failure':
            pass
        else:
            return
        #endif client_phase
    else:
        # (hopefully,) this is the initial phase; all others will be part of a POST request

        SAPDate = fnSAPList(req, CountDate)['SAPDate']

        cntext = {
            'SAP_Updated_at': SAPDate,
            'CountDate': CountDate,
            }
        templt = 'rpt_CountWkshtLoc.html'
    #endif req.method = 'POST'

    return render(req, templt, cntext)
    pass
