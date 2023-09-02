import datetime
import json
import os, uuid
import subprocess, signal
from django import forms
from django.conf import settings as django_settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models.query import QuerySet
from django.http import HttpRequest, HttpResponse
from django import http
from django.shortcuts import render
from django.views.generic import ListView
from django_q.tasks import async_task
from typing import *
from cMenu.models import getcParm
from cMenu.utils import makebool, isDate, WrapInQuotes, calvindate, ExcelWorkbook_fileext, Excelfile_fromqs
from mathematical_expressions_parser.eval import evaluate
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH
from WICS.models import ActualCounts, MaterialList, CountSchedule, Organizations, UploadSAPResults
from WICS.models_async_comm import async_comm, set_async_comm_state
from WICS.procs_SAP import fnSAPList

_MAX_LISTRECS = 5000

##############################################################
##############################################################
##############################################################

##### the suite of procs to support fnUploadCountSpreadsheet


def cleanupfld(fld, val, CountSprshtDateEpoch = WINDOWS_EPOCH):
    """
    fld is the name of the field in the ActualCount or MaterialList table
    val is the value to be cleaned for insertion into the fld
    Returns  {'usefld':usefld, 'cleanval': cleanval}
        usefld is a boolean indicating that val could/not be cleaned to the correct type
        cleanval is val in the correct type (if usefld==True)
    """
    cleanval = None

    if   fld == 'CountDate': 
        if isinstance(val,(calvindate, datetime.date, datetime.datetime)):
            usefld = True
            cleanval = calvindate(val).as_datetime()
        elif isinstance(val,int):
            usefld = True
            cleanval = from_excel(val,CountSprshtDateEpoch)
        else:
            usefld = isDate(val) 
            if (usefld != False):
                cleanval = calvindate(usefld).as_datetime()
                usefld = True
    elif fld in \
        ['CTD_QTY_Expr', 
            ]:
        if isinstance(val,str):
            if val[0] == '=':
                val = val[1:]
        try:
            v = evaluate(str(val))
        except (SyntaxError, NameError, TypeError, ZeroDivisionError):
            v = "-- INVALID --"
        usefld = (v!="-- INVALID --")
        cleanval = str(val) if (v != "--INVALID--") else None
    elif fld in \
        ['org_id', 
            'LocationOnly',
            'FLAG_PossiblyNotRecieved', 
            'FLAG_MovementDuringCount',
            ]:
        try:
            cleanval = int(val)
            usefld = True
        except:
            usefld = False
    elif fld in \
        ['Material', 
            'Counter', 
            'LOCATION', 
            'Notes', 
            'TypicalContainerQty', 
            'TypicalPalletQty',
            'PKGID_Desc',	
            'TAGQTY',
            ]:
        usefld = (val is not None)
        if usefld: cleanval = str(val)
    else:
        usefld = True
        cleanval = val
    
    return {'usefld':usefld, 'cleanval': cleanval}
#end def cleanupfld

def proc_UpActCountSprsheet_00InitUpld(reqid):
    acomm = set_async_comm_state(
        reqid,
        processname = 'Upload Counts',
        statecode = 'reading-spreadsht-init',
        statetext = 'Initializing',
        new_async=True,
        )
    UploadSAPResults.objects.all().delete()

def proc_UpActCountSprsheet_00CopySpreadsheet(req, reqid):
    acomm = set_async_comm_state(
        reqid,
        statecode = 'uploading-sprsht',
        statetext = 'Uploading Spreadsheet',
        )

    # save the file so we can open it as an excel file
    CountSprshtFile = req.FILES['CEFile']
    svdir = getcParm('SAP-FILELOC')
    fName = svdir+"tmpCE"+str(uuid.uuid4())+ExcelWorkbook_fileext
    with open(fName, "wb") as destination:
        for chunk in CountSprshtFile.chunks():
            destination.write(chunk)
    
    return fName

def proc_UpActCountSprsheet_01ReadSheet(reqid, fName):
    acomm = set_async_comm_state(
        reqid,
        statecode = 'rdng-sprsht',
        statetext = 'Reading Spreadsheet',
        )

    NOTdbFld_flags = ['**NOTdbFld**',]

    wb = load_workbook(filename=fName, read_only=True)
    CountSprshtDateEpoch = wb.epoch
    if 'Counts' in wb:
        ws = wb['Counts']
    else:
        acomm = set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = 'This workbook does not contain a sheet named Counts in the correct format',
            result = 'FAIL - no Counts sheet',
            )
        wb.close()
        os.remove(fName)
        return
    #endif 'Counts' in wb

    SprshtcolmnNames = ws[1]
    SprshtREQUIREDFLDS = ['Material','CountDate','Counter','LOCATION']     
        # LocationOnly/CTD_QTY_Expr handled separately since at least one must be present and both can be
    SprshtcolmnMap = {}
    Sprsht_SSName_TableName_map = {
            'CountDate': 'CountDate',
            'Counter': 'Counter',
            'LOCATION': 'LOCATION',
            'org_id': 'org_id',
            'Material': 'Material',
            'LocationOnly': 'LocationOnly',
            'CTD_QTY_Expr': 'CTD_QTY_Expr',
            'Typ Cntner Qty': 'TypicalContainerQty',
            'Typ Plt Qty': 'TypicalPalletQty',
            'Notes': 'Notes',
            'PKGID_Desc': 'PKGID_Desc',
            'TAGQTY': 'TAGQTY',
            'Poss Not Rcvd': 'FLAG_PossiblyNotRecieved',
            'Mvmt Dur Ct': 'FLAG_MovementDuringCount',
            'WICSignore': NOTdbFld_flags[0],
            }
    for col in SprshtcolmnNames:
        if col.value in Sprsht_SSName_TableName_map:
            colkey = Sprsht_SSName_TableName_map[col.value]
            # has this col.value already been mapped?
            if (colkey in SprshtcolmnMap and SprshtcolmnMap[colkey] is not None):
                # yes, that's a problem
                set_async_comm_state(
                    reqid,
                    statecode = 'fatalerr',
                    statetext = f'SAP Spreadsheet has bad header row - More than one column named {col.value}.  See Calvin to fix this.',
                    result = 'FAIL - bad spreadsheet',
                    )
                wb.close()
                os.remove(fName)
                return
            else:
                SprshtcolmnMap[colkey] = col.column - 1
            # endif previously mapped
        #endif col.value in SAP_SSName_TableName_map
    #endfor col in SAPcolmnNames

    HeaderGood = all([(reqFld in SprshtcolmnMap) for reqFld in SprshtREQUIREDFLDS])
    if not HeaderGood:
        MissingRequiredFields = [reqFld for reqFld in SprshtREQUIREDFLDS if reqFld not in SprshtcolmnMap]
        set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = f'Counts worksheet has bad header row - missing columns {MissingRequiredFields}.  See Calvin to fix this.',
            result = 'FAIL - bad spreadsheet',
            )
        wb.close()
        os.remove(fName)
        return

    SprshtRowNum=1
    nRowsAdded = 0
    nRowsNoMaterial = 0
    nRowsErrors = 0
    ABSMAX_COUNT_ROWS = 5000
    MAX_COUNT_ROWS = min(ABSMAX_COUNT_ROWS,ws.max_row)

    for row in ws.iter_rows(min_row=SprshtRowNum+1, max_row=MAX_COUNT_ROWS, values_only=True):
        SprshtRowNum += 1
        if SprshtRowNum % 100 == 0:
            set_async_comm_state(
                reqid,
                statecode = 'rdng-sprsht',
                statetext = f'Reading Spreadsheet ... record {SprshtRowNum} of {ws.max_row}',
                )

        ignoreline = ( ( (NOTdbFld_flags[0] in SprshtcolmnMap) and (row[SprshtcolmnMap[NOTdbFld_flags[0]]]) )
                       or (row[SprshtcolmnMap['Material']] is None)
                    )
        if not ignoreline:
            matlnum = cleanupfld('Material', row[SprshtcolmnMap['Material']])['cleanval']
            # if no org given, check that Material unique.
            if Sprsht_SSName_TableName_map['org_id'] not in SprshtcolmnMap:
                spshtorg = None
            else:
                spshtorg = cleanupfld('org_id', row[SprshtcolmnMap['org_id']])['cleanval']
            matlorglist = MaterialList.objects.filter(Material=matlnum).values_list('org_id', flat=True)
            MatlKount = len(matlorglist)
            MatObj = None
            err_already_handled = False
            if MatlKount == 1:
                MatObj = MaterialList.objects.get(Material=matlnum)
                spshtorg = MatObj.org_id
            if MatlKount > 1:
                if spshtorg is None:
                    UploadSAPResults(
                        errState = 'error',
                        errmsg = f"{matlnum} in multiple org_id's {tuple(matlorglist)}, but no org_id given", 
                        rowNum = SprshtRowNum
                        ).save()
                    nRowsErrors += 1
                    err_already_handled = True
                elif spshtorg in matlorglist:
                    MatObj = MaterialList.objects.get(org_id=spshtorg, Material=matlnum)
                else:
                    UploadSAPResults(
                        errState = 'error',
                        errmsg = f"{matlnum} in in multiple org_id's {tuple(matlorglist)}, but org_id given ({spshtorg}) is not one of them", 
                        rowNum = SprshtRowNum
                        ).save()
                    nRowsErrors += 1
                    err_already_handled = True
                #endif spshtorg is None
            #endif MatKount > 1

            if not MatObj:
                if not err_already_handled:
                    nRowsErrors += 1
                    UploadSAPResults(
                        errState = 'error',
                        errmsg = f'either {matlnum} does not exist in MaterialList or incorrect org_id ({str(spshtorg)}) given', 
                        rowNum = SprshtRowNum
                        ).save()
            else:
                rowErrs = False
                requiredFields = {reqFld: False for reqFld in SprshtREQUIREDFLDS}
                requiredFields['Both LocationOnly and CTD_QTY'] = False

                MatChanged = False
                SRec = ActualCounts()
                for fldName, colNum in SprshtcolmnMap.items():
                    if fldName in NOTdbFld_flags: continue
                    # check/correct problematic data types
                    usefld, V = cleanupfld(fldName, row[colNum], CountSprshtDateEpoch=CountSprshtDateEpoch).values()
                    if (V is not None):
                        if usefld:
                            if   fldName == 'CountDate': 
                                setattr(SRec, fldName, V) 
                                requiredFields['CountDate'] = True
                            elif fldName == 'Material': 
                                setattr(SRec, fldName, MatObj)
                                requiredFields['Material'] = True
                            elif fldName == 'Counter': 
                                setattr(SRec, fldName, V)
                                requiredFields['Counter'] = True
                            elif fldName == 'LOCATION': 
                                setattr(SRec, fldName, V)
                                requiredFields['LOCATION'] = True
                            elif fldName == 'LocationOnly': 
                                setattr(SRec, fldName, makebool(V))
                                requiredFields['Both LocationOnly and CTD_QTY'] = True
                            elif fldName == 'CTD_QTY_Expr': 
                                setattr(SRec, fldName, V)
                                requiredFields['Both LocationOnly and CTD_QTY'] = True
                            elif fldName == 'TypicalContainerQty' \
                            or fldName == 'TypicalPalletQty':
                                if V == '' or V == None: V = 0
                                if V != 0 and V != getattr(MatObj,fldName,0): 
                                    setattr(MatObj, fldName, V)
                                    MatChanged = True
                            else:
                                if hasattr(SRec, fldName): setattr(SRec, fldName, V)
                            # endif fldname
                        else:
                            if fldName!='CTD_QTY_Expr':
                                # we have to suspend judgement on CTD_QTY_Expr until last, because this could be a LocationOnly count
                                rowErrs = True
                                UploadSAPResults(
                                    errState = 'error',
                                    errmsg = f'{str(V)} is invalid for {fldName}', 
                                    rowNum = SprshtRowNum
                                    ).save()
                        #endif usefld
                    #endif (V is not None)
                # for each column
                
                # now we determine if one of LocationOnly or CTD_QTY was given
                if not requiredFields['Both LocationOnly and CTD_QTY']:
                    fldName = 'CTD_QTY_Expr'
                    V = row[SprshtcolmnMap[fldName]]
                    rowErrs = True
                    UploadSAPResults(
                        errState = 'error',
                        errmsg = f'record is not marked LocationOnly and {str(V)} is invalid for {fldName}', 
                        rowNum = SprshtRowNum
                        ).save()

                # are all required fields present?
                AllRequiredPresent = True
                for keyname, Prsnt in requiredFields.items():
                    AllRequiredPresent = AllRequiredPresent and Prsnt
                    if not Prsnt:
                        rowErrs = True
                        UploadSAPResults(
                            errState = 'error',
                            errmsg = f'{keyname} missing', 
                            rowNum = SprshtRowNum
                            ).save()

                if not rowErrs:
                    SRec.save()
                    if MatChanged: MatObj.save()
                    # res = {'error': False, 'rowNum':SprshtRowNum, 'TypicalQty':MatChanged, 'MaterialNum': str(MatObj) }
                    # res.update(SRec)      # tack the new record (along with its new pk) onto res
                    # UplResults.append(res)
                    resultString = str(SRec)
                    resultString += ' / LOCATION ONLY'  if SRec.LocationOnly else f' / Qty= {SRec.CTD_QTY_Expr}'
                    resultString += ' (Typ Cont Qty/Typ Plt Qty also changed)' if MatChanged else ''
                    UploadSAPResults(
                        errState = 'success',
                        errmsg = resultString, 
                        rowNum = SprshtRowNum
                        ).save()
                    nRowsAdded += 1
                else:
                    nRowsErrors += 1
                #endif not rowErrs
            # endif MatObj/not MatObj
        else:
            nRowsNoMaterial += 1
        #endif not ignoreline
    # endfor row in ws.iter_rows

    if SprshtRowNum >= ABSMAX_COUNT_ROWS:
        UploadSAPResults(
            errState = 'maxrows',
            errmsg = f'Data in spreadsheet rows {ABSMAX_COUNT_ROWS+1} and beyond are being ignored.', 
            rowNum = -1
            ).save()
    UploadSAPResults(
        errState = 'nRowsTotal',
        errmsg = '', 
        rowNum = SprshtRowNum
        ).save()
    UploadSAPResults(
        errState = 'nRowsAdded',
        errmsg = '', 
        rowNum = nRowsAdded
        ).save()
    UploadSAPResults(
        errState = 'nRowsErrors',
        errmsg = '', 
        rowNum = nRowsErrors
        ).save()
    UploadSAPResults(
        errState = 'nRowsIgnored',
        errmsg = '', 
        rowNum = nRowsNoMaterial
        ).save()

    # close and kill temp files
    wb.close()
    os.remove(fName)
def done_UpActCountSprsheet_01ReadSheet(t):
    reqid = t.args[0]
    statecode = async_comm.objects.get(pk=reqid).statecode
    if statecode != 'fatalerr':
        set_async_comm_state(
            reqid,
            statecode = 'done-rdng-sprsht',
            statetext = f'Finished Reading Spreadsheet',
            )
        proc_UpActCountSprsheet_99_FinalProc(reqid)
    #endif stateocde != 'fatalerr'

def proc_UpActCountSprsheet_99_FinalProc(reqid):
    set_async_comm_state(
        reqid,
        statecode = 'done',
        statetext = 'Finished Processing Spreadsheet',
        )

def proc_UpActCountSprsheet_99_Cleanup(reqid):
    # also kill reqid, acomm, qcluster process
    async_comm.objects.filter(pk=reqid).delete()
    os.kill(int(reqid), signal.SIGTERM)
    os.kill(int(reqid), signal.SIGKILL)

    # delete the temporary table
    UploadSAPResults.objects.all().delete()

@login_required
def fnUploadActCountSprsht(req):

    client_phase = req.POST['phase'] if 'phase' in req.POST else None
    reqid = req.COOKIES['reqid'] if 'reqid' in req.COOKIES else None

    if req.method == 'POST':
        # check any mandatory commits here and change status code

        if client_phase=='init-upl':
            retinfo = HttpResponse()

            # start django_q broker
            reqid = subprocess.Popen(
                ['python', f'{django_settings.BASE_DIR}/manage.py', 'qcluster']
            ).pid
            retinfo.set_cookie('reqid',str(reqid))
            proc_UpActCountSprsheet_00InitUpld(reqid)

            # save the file so we can open it as an excel file
            fName = proc_UpActCountSprsheet_00CopySpreadsheet(req, reqid)

            task01 = async_task(proc_UpActCountSprsheet_01ReadSheet, reqid, fName, hook=done_UpActCountSprsheet_01ReadSheet)

            acomm_fake = {
                'statecode': 'starting',
                'statetext': 'Count Spreadsheet Upload Starting',
                }
            retinfo.write(json.dumps(acomm_fake))
            return retinfo
        elif client_phase=='waiting':
            retinfo = HttpResponse()

            acomm = async_comm.objects.values().get(pk=reqid)    # something's very wrong if this doesn't exist
            stcode = acomm['statecode']
            if stcode == 'fatalerr':
                pass
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='wantresults':
            if UploadSAPResults.objects.filter(errState = 'nRowsTotal').exists(): SprshtRowNum = UploadSAPResults.objects.filter(errState = 'nRowsTotal')[0].rowNum
            else: SprshtRowNum = 0
            if UploadSAPResults.objects.filter(errState = 'nRowsAdded').exists(): nRowsAdded = UploadSAPResults.objects.filter(errState = 'nRowsAdded')[0].rowNum
            else: nRowsAdded = 0
            if UploadSAPResults.objects.filter(errState = 'nRowsErrors').exists(): nRowsErrors = UploadSAPResults.objects.filter(errState = 'nRowsErrors')[0].rowNum
            else: nRowsErrors = 0
            if UploadSAPResults.objects.filter(errState = 'nRowsIgnored').exists(): nRowsNoMaterial = UploadSAPResults.objects.filter(errState = 'nRowsIgnored')[0].rowNum
            else: nRowsNoMaterial = 0
            UplResults = UploadSAPResults.objects.exclude(errState__in = ['nRowsAdded','nRowsTotal','nRowsErrors','nRowsIgnored']).order_by('rowNum')
            cntext = {'UplResults':UplResults, 
                    'ResultStats': {
                            'nRowsRead': SprshtRowNum - 1,      
                                # -1 because header doesn't count
                            'nRowsAdded': nRowsAdded ,
                            'nRowsNoMaterial': nRowsNoMaterial,
                            'nRowsErrors': nRowsErrors,
                        },
                    }
            templt = 'frm_uploadCountEntry_Success.html'
            return render(req, templt, cntext)
        elif client_phase=='resultspresented':
            proc_UpActCountSprsheet_99_Cleanup(reqid)
            retinfo = HttpResponse()
            retinfo.delete_cookie('reqid')

            return retinfo
        else:
            return
        #endif client_phase

    else:   # req.method != 'POST'
        cntext = {
                }
        templt = 'frm_UploadCountEntrySprdsht.html'
    #endif req.method

    return render(req, templt, cntext)

#####################################################################
#####################################################################
#####################################################################

class ActualCountListForm(LoginRequiredMixin, ListView):
    ordering = ['-CountDate', 'Material']
    context_object_name = 'ActCtList'
    template_name = 'frm_ActualCountList.html'
    # http_method_names = ['get', 'post', ]   #'put', 'patch', 'delete', 'head', 'options', 'trace']
    
    def setup(self, req: HttpRequest, *args: Any, **kwargs: Any) -> None:
        self._user = req.user
        return super().setup(req, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Any]:
        return ActualCounts.objects.all().order_by(*self.ordering)[:_MAX_LISTRECS]
        # return super().get_queryset()

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        cntext = super().get_context_data(**kwargs)

        return cntext

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    # def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return super().get(request, *args, **kwargs)

    def post(self, req: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        # is there a request to copy a count?
        if ('copyCountFromid' in req.POST) and ('copyCountToDate' in req.POST):
            copyCountFromid = req.POST['copyCountFromid']
            copyCountToDate = req.POST['copyCountToDate']
            copyCountRec = ActualCounts.objects.get(pk=copyCountFromid)
            copyCountRec.pk = None
            copyCountRec.CountDate = copyCountToDate
            copyCountRec.save()
        # is there a request to delete a count
        if ('deleteCountid' in req.POST) :
            deleteCountid = req.POST['deleteCountid']
            ActualCounts.objects.filter(pk=deleteCountid).delete()
        
        return http.HttpResponseNotModified()




#####################################################################
#####################################################################
#####################################################################

@login_required
def fnCountSummaryReqRpt(req, passedCountDate='CURRENT_DATE'):
    return fnCountSummaryRpt(req, passedCountDate, Rptvariation='REQ')
@login_required
def fnCountSummaryRpt (req, passedCountDate='CURRENT_DATE', Rptvariation=None):

    # get the SAP data
    dtobj_pDate = isDate(passedCountDate)
    if not dtobj_pDate: dtobj_pDate = calvindate().as_datetime()
    SAP_SOH = fnSAPList(dtobj_pDate)

    # prep Excel_qdict.  It's up here so that the functions below have access to it
    Excel_qdict = []
    
    def CreateOutputRows(raw_qs, Eval_CTDQTY=True):
        def SummaryLine(lastrow):
            # summarize last Matl
            # total SAP Numbers
            SAPTot = 0
            outputline = dict()
            outputline['type'] = 'Summary'
            outputline['SAPNum'] = []
            for SAProw in SAP_SOH['SAPTable'].filter(Material_id=lastrow['Material_id']):
                outputline['SAPNum'].append((SAProw.StorageLocation, SAProw.Amount, SAProw.BaseUnitofMeasure))
                SAPTot += SAProw.Amount*SAProw.mult
            outputline['TypicalContainerQty'] = lastrow['TypicalContainerQty']
            outputline['TypicalPalletQty'] = lastrow['TypicalPalletQty']
            outputline['OrgName'] = lastrow['OrgName']
            outputline['Material'] = lastrow['Material']
            outputline['Material_id'] = lastrow['Material_id']
            outputline['Description'] = lastrow['Description']
            outputline['SchedCounter'] = lastrow['SchedCounter']
            outputline['Counters'] = lastrow['Counters']
            outputline['Requestor'] = lastrow['Requestor']
            outputline['RequestFilled'] = lastrow['RequestFilled']
            outputline['PartType'] = lastrow['PartType']
            outputline['CountTotal'] = lastrow['TotalCounted']
            outputline['SAPTotal'] = SAPTot
            outputline['Diff'] = lastrow['TotalCounted'] - SAPTot
            divsr = 1
            if lastrow['TotalCounted']!=0 or SAPTot!=0: divsr = max(lastrow['TotalCounted'], SAPTot)
            outputline['Accuracy'] = min(lastrow['TotalCounted'], SAPTot) / divsr * 100
            outputline['ReasonScheduled'] = lastrow['ReasonScheduled']
            outputline['SchedNotes'] = lastrow['SchedNotes']
            outputline['MatlNotes'] = lastrow['MatlNotes']
            #outputrows.append(outputline)

            return outputline
        # end def SummaryLine

        def CreateLastrow(rawrow):
            lastrow = dict()
            lastrow['OrgName'] = rawrow.OrgName
            lastrow['Material'] = rawrow.Matl_PartNum
            lastrow['Material_id'] = rawrow.matl_id
            lastrow['Description'] = rawrow.Description
            lastrow['SchedCounter'] = rawrow.cs_Counter
            lastrow['Counters'] = rawrow.ac_Counter if rawrow.ac_Counter is not None else ''
            lastrow['Requestor'] = rawrow.Requestor
            lastrow['RequestFilled'] = rawrow.RequestFilled
            lastrow['PartType'] = rawrow.PartType
            lastrow['TotalCounted'] = 0
            lastrow['SchedNotes'] = rawrow.cs_Notes
            lastrow['TypicalContainerQty'] = rawrow.TypicalContainerQty
            lastrow['TypicalPalletQty'] = rawrow.TypicalPalletQty
            lastrow['MatlNotes'] = rawrow.mtl_Notes
            lastrow['ReasonScheduled'] = rawrow.cs_ReasonScheduled

            return lastrow
        # end def CreateLastRow

        def DetailLine(rawrow, Eval_CTDQTY=True):
            outputline = dict()
            outputline['type'] = 'Detail'
            outputline['CycCtID'] = rawrow.ac_CycCtID
            outputline['Material'] = rawrow.Matl_PartNum
            outputline['Material_id'] = rawrow.matl_id
            outputline['org_id'] = rawrow.org_id
            outputline['orgName'] = rawrow.OrgName
            outputline['ActCounter'] = rawrow.ac_Counter
            if rawrow.ac_Counter is not None and rawrow.ac_Counter not in lastrow['Counters']:
                lastrow['Counters'] += ', ' + rawrow.ac_Counter
            outputline['LOCATION'] = rawrow.ac_LOCATION
            outputline['PKGID'] = rawrow.ac_PKGID_Desc
            outputline['TAGQTY'] = rawrow.ac_TAGQTY
            outputline['PossNotRec'] = rawrow.FLAG_PossiblyNotRecieved
            outputline['MovDurCt'] = rawrow.FLAG_MovementDuringCount
            outputline['CTD_QTY_Expr'] = rawrow.ac_CTD_QTY_Expr
            if Eval_CTDQTY:
                try:
                    outputline['CTD_QTY_Eval'] = evaluate(rawrow.ac_CTD_QTY_Expr)
                    # do next line at caller
                    # lastrow['TotalCounted'] += outputline['CTD_QTY_Eval']
                except:
                    # Exception('bad expression:'+rawrow.ac_CTD_QTY_Expr)
                    outputline['CTD_QTY_Eval'] = "????"
            else:
                outputline['CTD_QTY_Eval'] = "----"
            outputline['ActCountNotes'] = rawrow.ac_Notes
            # outputrows.append(outputline)

            return outputline
        #end def DetailLine
        
        outputrows = []
        lastrow = {'Material_id': None}
        for rawrow in raw_qs:
            if rawrow.matl_id != lastrow['Material_id']:     # new Matl
                if outputrows:
                    SmLine = SummaryLine(lastrow)
                    outputrows.append(SmLine)
                    Excel_qdict.append(
                        {key:SmLine[key] 
                          for key in ['OrgName','Material','PartType','Description','CountTotal','SAPTotal','Diff','Accuracy','Counters']
                        })
                # no else -  if outputrows is empty, this is the first row, so keep going

                # this new material is now the "old" one; save values for when it switches, and we do the above block
                # this whole block becomes
                lastrow = CreateLastrow(rawrow)
            #endif

            # process this row
            outputline = DetailLine(rawrow, Eval_CTDQTY)
            outputrows.append(outputline)
            if isinstance(outputline['CTD_QTY_Eval'],(int,float)): lastrow['TotalCounted'] += outputline['CTD_QTY_Eval']
        # endfor
        # need to do the summary on the last row
        if outputrows:
            # summarize last Matl
            SmLine = SummaryLine(lastrow)
            outputrows.append(SmLine)
            Excel_qdict.append(
                {key:SmLine[key] 
                    for key in ['OrgName','Material','PartType','Description','CountTotal','SAPTotal','Diff','Accuracy','Counters']
                })
        
        return outputrows
    #end def CreateOutputRows
        
    ### main body of fnCountSummaryRpt

    SummaryReport = []

    fldlist = "0 as id, cs.id as cs_id, cs.CountDate as cs_CountDate , cs.Counter as cs_Counter" \
        ", cs.Priority as cs_Priority, cs.ReasonScheduled as cs_ReasonScheduled" \
        ", cs.Requestor, cs.RequestFilled" \
        ", cs.Notes as cs_Notes" \
        ", ac.id as ac_id, ac.CountDate as ac_CountDate, ac.CycCtID as ac_CycCtID, ac.Counter as ac_Counter" \
        ", ac.LocationOnly as ac_LocationOnly, ac.CTD_QTY_Expr as ac_CTD_QTY_Expr" \
        ", ac.LOCATION as ac_LOCATION, ac.PKGID_Desc as ac_PKGID_Desc, ac.TAGQTY as ac_TAGQTY" \
        ", ac.FLAG_PossiblyNotRecieved, ac.FLAG_MovementDuringCount, ac.Notes as ac_Notes" \
        ", mtl.id as matl_id, mtl.org_id, mtl.OrgName" \
        ", mtl.Material_org as Matl_PartNum, mtl.PartType as PartType" \
        ", mtl.Description, mtl.TypicalContainerQty, mtl.TypicalPalletQty, mtl.Notes as mtl_Notes"
    if isDate(passedCountDate): datestr = WrapInQuotes(passedCountDate,"'","'")
    else: datestr = passedCountDate
    date_condition = '(ac.CountDate = ' + datestr + ' OR cs.CountDate = ' + datestr + ') '
    order_by = 'Matl_PartNum'

    #VIEW_Material_sql = "(select MATL.id AS id, MATL.org_id AS org_id, MATL.Material AS Material, MATL.Description AS Description,"
    #VIEW_Material_sql += "  MATL.Notes AS Notes, MATL.PartType_id AS PartType_id,"
    #VIEW_Material_sql += "  MATL.TypicalContainerQty AS TypicalContainerQty, MATL.TypicalPalletQty AS TypicalPalletQty,"
    #VIEW_Material_sql += "  PTYPE.WhsePartType AS PartType, ORG.orgname AS OrgName,"
    #VIEW_Material_sql += "  if((exists "
    #VIEW_Material_sql += "         (select * from WICS_materiallist numdups where ((numdups.Material = MATL.Material) and (numdups.org_id <> MATL.org_id)))) , "
    #VIEW_Material_sql += "       concat(MATL.Material, ' (', ORG.orgname, ')') , "
    #VIEW_Material_sql += "       MATL.Material "
    #VIEW_Material_sql += "     ) AS Material_org "
    #VIEW_Material_sql += "from "
    #VIEW_Material_sql += "  ((WICS_materiallist MATL join WICS_organizations ORG on (MATL.org_id = ORG.id)) left join WICS_whseparttypes `PTYPE` on MATL.PartType_id = `PTYPE`.id) "
    #VIEW_Material_sql += " ) mtl "
    VIEW_Material_sql = "VIEW_materials mtl "

    for org in Organizations.objects.all():
        # group by org_id
        org_condition = '(mtl.org_id = ' + str(org.pk) + ')'

        A_Sched_Ctd_from = 'WICS_countschedule cs INNER JOIN ' + VIEW_Material_sql 
        A_Sched_Ctd_from += ' INNER JOIN (SELECT * FROM WICS_actualcounts WHERE not LocationOnly) ac '    
        A_Sched_Ctd_joinon = ' cs.CountDate=ac.CountDate AND cs.Material_id=ac.Material_id AND ac.Material_id=mtl.id'    
        A_Sched_Ctd_where = ''
        if Rptvariation == 'REQ':
            if A_Sched_Ctd_where:  A_Sched_Ctd_where += ' AND ' 
            A_Sched_Ctd_where += 'Requestor IS NOT NULL'
        A_Sched_Ctd_sql = 'SELECT ' + fldlist + \
            ' FROM ' + A_Sched_Ctd_from + \
            ' ON ' + A_Sched_Ctd_joinon + \
            ' WHERE NOT ac.LocationOnly AND ' + date_condition + ' AND ' + org_condition
        if A_Sched_Ctd_where:
            A_Sched_Ctd_sql += ' AND ' + A_Sched_Ctd_where
        A_Sched_Ctd_sql += ' ORDER BY ' + order_by
        A_Sched_Ctd_qs = CountSchedule.objects.raw(A_Sched_Ctd_sql)
        # build display lines
        ttl = 'Scheduled and Counted'
        if Rptvariation == 'REQ':
            ttl = 'Requested and Counted'            
        SummaryReport.append({
                    'org':org,
                    'Title':ttl,
                    'outputrows': CreateOutputRows(A_Sched_Ctd_qs)
                    })

        if Rptvariation is None:
            B_UnSched_Ctd_from = 'WICS_countschedule cs RIGHT JOIN' \
                ' ((SELECT * FROM WICS_actualcounts WHERE not LocationOnly) ac INNER JOIN ' + VIEW_Material_sql + ' ON ac.Material_id=mtl.id)'
            B_UnSched_Ctd_joinon = 'cs.CountDate=ac.CountDate AND cs.Material_id=ac.Material_id'
            B_UnSched_Ctd_where = '(cs.id IS NULL)'
            B_UnSched_Ctd_sql = 'SELECT ' + fldlist + ' ' + \
                ' FROM ' + B_UnSched_Ctd_from + \
                ' ON ' + B_UnSched_Ctd_joinon + \
                ' WHERE NOT ac.LocationOnly AND ' + date_condition + ' AND ' + org_condition + \
                ' AND ' + B_UnSched_Ctd_where + \
                ' ORDER BY ' + order_by
            B_UnSched_Ctd_qs = CountSchedule.objects.raw(B_UnSched_Ctd_sql)
            SummaryReport.append({
                        'org':org,
                        'Title':'UnScheduled',
                        'outputrows': CreateOutputRows(B_UnSched_Ctd_qs)
                        })
        
        C_Sched_NotCtd_Ctd_from = '(WICS_countschedule cs INNER JOIN ' + VIEW_Material_sql + ' ON cs.Material_id=mtl.id)' 
        C_Sched_NotCtd_Ctd_from += ' LEFT JOIN (SELECT * FROM WICS_actualcounts WHERE not LocationOnly) ac '    
        C_Sched_NotCtd_Ctd_joinon = 'cs.CountDate=ac.CountDate AND cs.Material_id=ac.Material_id'
        C_Sched_NotCtd_Ctd_where = '(ac.id IS NULL)'
        if Rptvariation == 'REQ':
            if C_Sched_NotCtd_Ctd_where:  C_Sched_NotCtd_Ctd_where += ' AND ' 
            C_Sched_NotCtd_Ctd_where += '(Requestor IS NOT NULL)'
        C_Sched_NotCtd_Ctd_sql = 'SELECT ' + fldlist + ' ' + \
            ' FROM ' + C_Sched_NotCtd_Ctd_from + \
            ' ON ' + C_Sched_NotCtd_Ctd_joinon + \
            ' WHERE ' + date_condition + ' AND ' + org_condition
        if C_Sched_NotCtd_Ctd_where:
            C_Sched_NotCtd_Ctd_sql += ' AND ' + C_Sched_NotCtd_Ctd_where
        C_Sched_NotCtd_Ctd_sql += ' ORDER BY ' + order_by
        C_Sched_NotCtd_Ctd_qs = CountSchedule.objects.raw(C_Sched_NotCtd_Ctd_sql)
        ttl = 'Scheduled but Not Counted'
        if Rptvariation == 'REQ':
            ttl = 'Requested but Not Counted'            
        SummaryReport.append({
                    'org':org,
                    'Title':ttl,
                    'outputrows': CreateOutputRows(C_Sched_NotCtd_Ctd_qs, Eval_CTDQTY=False)
                    })
    
    AccuracyCutoff = { 
                'DANGER': float(getcParm('ACCURACY-DANGER')),
                'SUCCESS': float(getcParm('ACCURACY-SUCCESS')),
                'WARNING': float(getcParm('ACCURACY-WARNING')),
                }

    ExcelFileNamePrefix = "CountSummary "
    svdir = django_settings.STATIC_ROOT if django_settings.STATIC_ROOT is not None else django_settings.STATICFILES_DIRS[0]
    fName_base = '/tmpdl/'+ExcelFileNamePrefix + f'{dtobj_pDate:%Y-%m-%d}'
    fName = svdir + fName_base
    ExcelFileName = Excelfile_fromqs(Excel_qdict, fName)

    # display the form
    cntext = {
            'variation': Rptvariation,
            'CountDate': dtobj_pDate,
            'SAPDate': SAP_SOH['SAPDate'],
            'AccuracyCutoff': AccuracyCutoff,
            'SummaryReport': SummaryReport,
            'FilSavLoc': ExcelFileName,
            'ExcelFileName': fName_base+ExcelWorkbook_fileext, 
            }
    templt = 'rpt_CountSummary.html'
    return render(req, templt, cntext)


