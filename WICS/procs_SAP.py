import os, uuid, re as regex
import ast
import subprocess, signal
import json
from functools import partial
from django import forms
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.conf import settings as django_settings
from django.db import connection, connections, transaction
from django.db.models import Max, OuterRef, Subquery
from django.http import HttpRequest, HttpResponse #, HttpResponseNotModified
from django.shortcuts import render
from django_q.tasks import async_task
from openpyxl import load_workbook
import WICS.procs_SAP
from cMenu.models import getcParm
from cMenu.utils import calvindate, ExcelWorkbook_fileext
from cMenu.views import user_db
import WICS.globals
from WICS.models import SAP_SOHRecs, SAPPlants_org, UnitsOfMeasure, UploadSAPResults  #, VIEW_SAP
from WICS.models import WhsePartTypes, MaterialList, tmpMaterialListUpdate
from WICS.models_async_comm import async_comm, set_async_comm_state


def fnSAPExists(req:HttpRequest, reqDate:calvindate=calvindate().today()) -> bool:
    """
    returns true or false indicating if SAP_SOH data exists for reqDate
    """

    return SAP_SOHRecs.objects.using(user_db(req)).filter(uploaded_at=calvindate(reqDate).as_datetime()).exists()
def fnajaxSAPExists(req, reqDate=calvindate().today()):
    """
    returns true or false to an ajax caller indicating if SAP_SOH data exists for reqDate
    """

    retinfo = HttpResponse()

    retinfo.write(json.dumps(fnSAPExists(req, reqDate)))
    return retinfo


@login_required
def fnShowSAP(req, reqDate=calvindate().today()):

    reqDate = calvindate(reqDate).as_datetime()
    _myDtFmt = '%Y-%m-%d'

    SAP_tbl = fnSAPList(req, for_date=reqDate)
    SAPDatesRaw = SAP_SOHRecs.objects.using(user_db(req)).order_by('-uploaded_at').values('uploaded_at').distinct()
    SAPDates = []
    for D in SAPDatesRaw:
        SAPDates.append(D['uploaded_at'].strftime(_myDtFmt))


    cntext = {'reqDate': SAP_tbl['reqDate'],
            'SAPDateList': SAPDates,
            'SAPDate': SAP_tbl['SAPDate'].strftime(_myDtFmt),
            'SAPSet': SAP_tbl['SAPTable'],
            }
    templt = 'show_SAP_table.html'
    return render(req, templt, cntext)


####################################################################################
####################################################################################
####################################################################################

##### the suite of procs to support fnUploadSAP


class UploadSAPForm(forms.Form):
    uploaded_at = forms.DateField(widget=forms.widgets.DateInput())
    SAPFile = forms.FileField()

def proc_UpSAPSpreadsheet_00InitUpSAP(dbToUse, reqid):
    acomm = set_async_comm_state(
        dbToUse,
        reqid,
        processname = 'Upload SAP MM52',
        statecode = 'reading-spreadsht-init',
        statetext = 'Initializing',
        new_async=True,
        )
    UploadSAPResults.objects.using(dbToUse).all().delete()

def proc_UpSAPSpreadsheet_00CopyUpSAPSpreadsheet(req, reqid):
    dbToUse = user_db(req)
    acomm = set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'uploading-sprsht',
        statetext = 'Uploading Spreadsheet',
        )

    SAPFile = req.FILES['SAPFile']
    svdir = getcParm(req, 'SAP-FILELOC')
    fName = svdir+"tmpSAP"+str(uuid.uuid4())+ExcelWorkbook_fileext
    with open(fName, "wb") as destination:
        for chunk in SAPFile.chunks():
            destination.write(chunk)

    return fName

def proc_UpSAPSpreadsheet_01ReadSheet(dbToUse, reqid, fName, UplDate):
    acomm = set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'rdng-sprsht',
        statetext = 'Reading Spreadsheet',
        )

    nRowsAdded = 0
    SprshtRowNum = 0

    wb = load_workbook(filename=fName, read_only=True)
    ws = wb.active

    _SStName_Material = 'Material'
    _TblName_Material = 'MaterialPartNum'
    _SStName_Plant = 'Plant'
    _TblName_Plant = 'Plant'
    SAPcolmnNames = ws[1]
    SAPcolmnMap = {_TblName_Material: None, _TblName_Plant: None}
    SAP_SSName_TableName_map = {
            _SStName_Material: _TblName_Material,  # Material+org will translate to a Material_id
            _SStName_Plant: _TblName_Plant,
            'Material description': 'Description',
            'Material type': 'MaterialType',
            'Storage location': 'StorageLocation',
            'Base Unit of Measure': 'BaseUnitofMeasure',
            'Unrestricted': 'Amount',
            'Currency': 'Currency',
            'Value Unrestricted': 'ValueUnrestricted',
            'Special Stock': 'SpecialStock',
            'Blocked':'Blocked',
            'Value BlockedStock':'ValueBlocked',
            'Vendor':'Vendor',
            'Batch': 'Batch',
            }
    for col in SAPcolmnNames:
        if col.value in SAP_SSName_TableName_map:
            colkey = SAP_SSName_TableName_map[col.value]
            # has this col.value already been mapped?
            if (colkey in SAPcolmnMap and SAPcolmnMap[colkey] is not None):
                # yes, that's a problem
                set_async_comm_state(
                    dbToUse,
                    reqid,
                    statecode = 'fatalerr',
                    statetext = f'SAP Spreadsheet has bad header row - More than one column named {col.value}.  See Calvin to fix this.',
                    result = 'FAIL - bad spreadsheet',
                    )
                wb.close()
                os.remove(fName)
                return
            else:
                SAPcolmnMap[colkey] = col.column - 1
            # endif previously mapped
        #endif col.value in SAP_SSName_TableName_map
    #endfor col in SAPcolmnNames
    if (SAPcolmnMap[_TblName_Material] is None) or (SAPcolmnMap[_TblName_Plant] is None):   # or SAPcol['StorageLocation'] == None or SAPcol['Amount'] == None):
        set_async_comm_state(
            dbToUse,
            reqid,
            statecode = 'fatalerr',
            statetext = f'SAP Spreadsheet has bad header row - no {_SStName_Material} column or no {_SStName_Plant} column.  See Calvin to fix this.',
            result = 'FAIL - bad spreadsheet',
            )
        wb.close()
        os.remove(fName)
        return

    # if SAP SOH records exist for this date, kill them; only one set of SAP SOH records per day
    # (this was signed off on by user before coming here)
    SAP_SOHRecs.objects.using(dbToUse).filter(uploaded_at=UplDate).delete()

    numrows = ws.max_row
    SprshtRowNum = 1
    for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
        SprshtRowNum += 1
        if SprshtRowNum % 100 == 0:
            set_async_comm_state(
                dbToUse,
                reqid,
                statecode = 'rdng-sprsht',
                statetext = f'Reading Spreadsheet ... record {SprshtRowNum} of {numrows}<br><progress max="{numrows}" value="{SprshtRowNum}"></progress>',
                )

        if row[SAPcolmnMap[_TblName_Material]]==None: MatlNum = ''
        else: MatlNum = row[SAPcolmnMap[_TblName_Material]]
        if len(str(MatlNum)):
            _org = SAPPlants_org.objects.using(dbToUse).filter(SAPPlant=row[SAPcolmnMap[_TblName_Plant]])[0].org
            try:
                MatlRec = MaterialList.objects.using(dbToUse).get(org=_org,Material=MatlNum)
            except:
                MatlRec = None
            if not MatlRec:
                UploadSAPResults(
                    errState = 'error',
                    errmsg = f'either {MatlNum}  does not exist in MaterialList or incorrect Plant ({str(row[SAPcolmnMap[_TblName_Plant]])}) given',
                    rowNum = SprshtRowNum
                    ).save(using=dbToUse)
            else:
                SRec = SAP_SOHRecs(
                        org = _org,     # will be going away - or will it???
                        uploaded_at = UplDate,
                        Material = MatlRec
                        )
                for fldName, colNum in SAPcolmnMap.items():
                    if fldName == _TblName_Material:
                        pass    # not continue - we are preserving the incoming MaterialPartNum string
                    if row[colNum] is None: setval = ''
                    else: setval = row[colNum]
                    setattr(SRec, fldName, setval)
                SRec.save(using=dbToUse)
                nRowsAdded += 1
            # endif MatlRec
        # endif len(str(MatlNum))
    #endfor row in ws.iter_rows

    UploadSAPResults(
        errState = 'nRowsTotal',
        errmsg = '',
        rowNum = SprshtRowNum
        ).save(using=dbToUse)
    UploadSAPResults(
        errState = 'nRowsAdded',
        errmsg = '',
        rowNum = nRowsAdded
        ).save(using=dbToUse)

    # close and kill temp files
    wb.close()
    os.remove(fName)
def done_UpSAPSpreadsheet_01ReadSheet(t):
    dbToUse = t.args[0]
    reqid = t.args[1]
    statecode = async_comm.objects.using(dbToUse).get(pk=reqid).statecode
    if statecode != 'fatalerr':
        set_async_comm_state(
            dbToUse,
            reqid,
            statecode = 'done-rdng-sprsht',
            statetext = f'Finished Reading Spreadsheet',
            )
        proc_UpSAPSpreadsheet_99_FinalProc(dbToUse, reqid)
    #endif stateocde != 'fatalerr'

def proc_UpSAPSpreadsheet_99_FinalProc(dbToUse, reqid):
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'done',
        statetext = 'Finished Processing Spreadsheet',
        )

def proc_UpSAPSpreadsheet_99_Cleanup(dbToUse, reqid):
    # also kill reqid, acomm, qcluster process
    async_comm.objects.using(dbToUse).filter(pk=reqid).delete()

    try:
        os.kill(int(reqid), signal.SIGTERM)
    except AttributeError:
        pass
    try:
        os.kill(int(reqid), signal.SIGKILL)
    except AttributeError:
        pass

    # delete the temporary table
    UploadSAPResults.objects.using(dbToUse).all().delete()

@login_required
def fnUploadSAP(req):

    dbToUse = user_db(req)
    client_phase = req.POST['phase'] if 'phase' in req.POST else None
    reqid = req.COOKIES['reqid'] if 'reqid' in req.COOKIES else None
    UplDate = calvindate(req.POST['uploaded_at']).isoformat() if 'uploaded_at' in req.POST else calvindate()

    if req.method == 'POST':

        if client_phase=='init-upl':
            retinfo = HttpResponse()

            # start django_q broker
            reqid = subprocess.Popen(
                ['python', f'{django_settings.BASE_DIR}/manage.py', 'qcluster']
            ).pid
            retinfo.set_cookie('reqid',str(reqid))
            proc_UpSAPSpreadsheet_00InitUpSAP(dbToUse, reqid)

            # save the file so we can open it as an excel file
            fName = proc_UpSAPSpreadsheet_00CopyUpSAPSpreadsheet(req, reqid)

            task01 = async_task(proc_UpSAPSpreadsheet_01ReadSheet, dbToUse, reqid, fName, UplDate, hook=done_UpSAPSpreadsheet_01ReadSheet)

            acomm_fake = {
                'statecode': 'starting',
                'statetext': 'SAP MM60 Update Starting',
                }
            retinfo.write(json.dumps(acomm_fake))
            return retinfo
        elif client_phase=='waiting':
            retinfo = HttpResponse()

            acomm = async_comm.objects.using(dbToUse).values().get(pk=reqid)    # something's very wrong if this doesn't exist
            stcode = acomm['statecode']
            if stcode == 'fatalerr':
                pass
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='wantresults':
            if UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsAdded').exists():
                nRowsAdded = UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsAdded')[0].rowNum
            else:
                nRowsAdded = 0
            if UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsTotal').exists():
                nRowsTotal = UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsTotal')[0].rowNum
            else:
                nRowsTotal = 0
            UplResults = UploadSAPResults.objects.using(dbToUse).exclude(errState__in = ['nRowsAdded','nRowsTotal'])
            cntext = {
                'uploaded_at':UplDate,
                'nRows':nRowsAdded,
                'nRowsRead':nRowsTotal,
                'UplProblems': UplResults,
                    }
            templt = 'frm_upload_SAP_Success.html'
            return render(req, templt, cntext)
        elif client_phase=='resultspresented':
            proc_UpSAPSpreadsheet_99_Cleanup(dbToUse, reqid)
            retinfo = HttpResponse()
            retinfo.delete_cookie('reqid')

            return retinfo
        else:
            return
        #endif client_phase

    else:   # req.method != 'POST'
        LastSAPUpload = SAP_SOHRecs.objects.using(dbToUse).all().aggregate(LastSAPDate=Max('uploaded_at'))
        # .first().values('LastSAPDate')['LastSAPDate']

        form = UploadSAPForm()
        cntext = {'form': form,
                'LastSAPUploadDate': LastSAPUpload['LastSAPDate'],
                }
        templt = 'frm_upload_SAP.html'
    #endif  req.method == 'POST'

    return render(req, templt, cntext)


####################################################################################
####################################################################################
####################################################################################


# read the last SAP list before for_date into a list of SAP_SOHRecs
def fnSAPList(req:HttpRequest|User|str, for_date = calvindate().today(), matl = None) -> dict:
    """
    finally done!: allow matl to be a MaterialList object or an id
    matl is a Material (string, NOT object!), or list, tuple or queryset of Materials to list, or None if all records are to be listed
    the SAPDate returned is the last one prior or equal to for_date
    """
    _myDtFmt = '%Y-%m-%d %H:%M'

    dateObj = calvindate(for_date) #.as_datetime().date()

    db_to_use = 'default'
    if   isinstance(req, (HttpRequest, User)): 
        db_to_use = user_db(req)
    elif isinstance(req, str):
        db_to_use = req
    #endif isinstance(req)

    if SAP_SOHRecs.objects.using(db_to_use).exists():
        if SAP_SOHRecs.objects.using(db_to_use).filter(uploaded_at__lte=dateObj).exists():
            LatestSAPDate = SAP_SOHRecs.objects.using(db_to_use).filter(uploaded_at__lte=dateObj).latest().uploaded_at
        else:
            LatestSAPDate = SAP_SOHRecs.objects.using(db_to_use).earliest().uploaded_at
    else:
        LatestSAPDate = None
    SAPLatest = SAP_SOHRecs.objects.using(db_to_use)\
        .filter(uploaded_at=LatestSAPDate)\
        .annotate(mult=Subquery(UnitsOfMeasure.objects.filter(UOM=OuterRef('BaseUnitofMeasure')).values('Multiplier1')[:1]))\
        .order_by('Material__org', 'Material__Material', 'StorageLocation')

    SList = {'reqDate': for_date, 'SAPDate': LatestSAPDate, 'SAPTable':[]}

    if not matl:
        STable = SAPLatest
    else:
        if isinstance(matl,str):
            raise Exception('fnSAPList by Matl string is deprecated')
        elif isinstance(matl,MaterialList):  # handle case matl is a MaterialList instance here
            STable = SAPLatest.filter(Material=matl)
        elif isinstance(matl,int):  # handle case matl is a MaterialList id here
            STable = SAPLatest.filter(Material_id=matl)
        else:   # it better be an iterable!
            STable = SAPLatest.filter(Material__in=matl)

    # yea, building SList is sorta wasteful, but a lot of existing code depends on it
    # won't be changing it until a total revamp of WICS
    if not STable:
        SList['SAPDate'] = None
    SList['SAPTable'] = STable

    return SList


####################################################################################
####################################################################################
####################################################################################

##### the suite of procs to support fnUpdateMatlListfromSAP

def proc_MatlListSAPSprsheet_00InitUMLasync_comm(dbToUse, reqid, UpdateExistFldList):
    acomm = set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'rdng-sprsht-init',
        statetext = 'Initializing ...',
        new_async=True
        )
    set_async_comm_state(
        dbToUse, 
        f'{reqid}-UpdExstFldList',
        statecode = 'UpdateExistFldList',
        statetext = f'{UpdateExistFldList}',
        new_async=True
        )

def proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(req, reqid):
    dbToUse = user_db(req)
    acomm = set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'uploading-sprsht',
        statetext = 'Uploading Spreadsheet',
        )

    SAPFile = req.FILES['SAPFile']
    svdir = getcParm(req, 'SAP-FILELOC')
    fName = svdir+"tmpMatlList"+str(uuid.uuid4())+ExcelWorkbook_fileext
    with open(fName, "wb") as destination:
        for chunk in SAPFile.chunks():
            destination.write(chunk)

    return fName

def proc_MatlListSAPSprsheet_01ReadSpreadsheet(dbToUse, reqid, fName):
    acomm = set_async_comm_state(
        dbToUse, 
        reqid,
        statecode = 'rdng-sprsht',
        statetext = 'Reading Spreadsheet',
        )

    tmpMaterialListUpdate.objects.using(dbToUse).all().delete()

    wb = load_workbook(filename=fName, read_only=True)
    ws = wb.active
    SAPcolmnNames = ws[1]
    SAPcol = {'Plant':None,'Material': None}
    SAP_SSName_TableName_map = {
            'Material': 'Material',
            'Material description': 'Description',
            'Plant': 'Plant', 'Plnt': 'Plant',
            'Material type': 'SAPMaterialType',  'MTyp': 'SAPMaterialType',
            'Material Group': 'SAPMaterialGroup', 'Matl Group': 'SAPMaterialGroup',
            'Manufact.': 'SAPManuf', 
            'MPN': 'SAPMPN', 
            'ABC': 'SAPABC', 
            'Price': 'Price', 'Standard price': 'Price',
            'Price unit': 'PriceUnit', 'per': 'PriceUnit',
            'Currency':'Currency',
            }
    for col in SAPcolmnNames:
        if col.value in SAP_SSName_TableName_map:
            SAPcol[SAP_SSName_TableName_map[col.value]] = col.column - 1
    if (SAPcol['Material'] == None or SAPcol['Plant'] == None):
        set_async_comm_state(
            dbToUse,
            reqid,
            statecode = 'fatalerr',
            statetext = 'SAP Spreadsheet has bad header row. Plant and/or Material is missing.  See Calvin to fix this.',
            result = 'FAIL - bad spreadsheet',
            )

        wb.close()
        os.remove(fName)
        return

    numrows = ws.max_row
    nRows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nRows += 1
        if nRows % 100 == 0:
            set_async_comm_state(
                dbToUse,
                reqid,
                statecode = 'rdng-sprsht',
                statetext = f'Reading Spreadsheet ... record {nRows} of {numrows}<br><progress max="{numrows}" value="{nRows}"></progress>',
                )

        if row[SAPcol['Material']]==None: MatNum = ''
        else: MatNum = row[SAPcol['Material']]
        validTmpRec = False
        ## create a blank tmpMaterialListUpdate record,
        newrec = tmpMaterialListUpdate()
        if regex.match(".*[\n\t\xA0].*",MatNum):
            validTmpRec = True
            ## refuse to work with special chars embedded in the MatNum
            newrec.recStatus = 'err-MatlNum'
            newrec.errmsg = f'error: {MatNum!a} is an unusable part number. It contains invalid characters and cannot be added to WICS'
        elif len(str(MatNum)):
            validTmpRec = True
            _org = SAPPlants_org.objects.using(dbToUse).filter(SAPPlant=row[SAPcol['Plant']])[0].org        #TODO: handle empty table
            newrec.org = _org
        # endif invalid Material
        if validTmpRec:
            ## populate by looping through SAPcol,
            ## then save
            for dbColName, ssColNum in SAPcol.items():
                setattr(newrec,dbColName,row[ssColNum])
            newrec.save(using=dbToUse)
    # endfor

    wb.close()
    os.remove(fName)
def done_MatlListSAPSprsheet_01ReadSpreadsheet(t):
    dbToUse = t.args[0]
    reqid = t.args[1]
    statecode = async_comm.objects.using(dbToUse).get(pk=reqid).statecode
    #DOITNOW!!! handle not t.success, t.result
    if statecode != 'fatalerr':
        set_async_comm_state(
            dbToUse,
            reqid,
            statecode = 'done-rdng-sprsht',
            statetext = f'Finished Reading Spreadsheet',
            )
        proc_MatlListSAPSprsheet_02_identifyexistingMaterial(dbToUse, reqid)

def proc_MatlListSAPSprsheet_02_identifyexistingMaterial(dbToUse, reqid):
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'get-matl-link',
        statetext = f'Finding SAP MM60 Materials already in WICS Material List',
        )
    UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
    UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id, '
    UpdMaterialLinkSQL += "     WICS_tmpmateriallistupdate.recStatus = 'FOUND' "
    UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
    UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
    with connections[dbToUse].cursor() as cursor:
        cursor.execute(UpdMaterialLinkSQL)

    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'id-del-matl',
        statetext = f'Identifying WICS Materials no longer in SAP MM60 Materials',
        )
    MustKeepMatlsSelCond = ''
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT tmucopy.MaterialLink_id AS Material_id FROM WICS_tmpmateriallistupdate tmucopy WHERE tmucopy.MaterialLink_id IS NOT NULL)'
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_actualcounts)'
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_countschedule)'
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_sap_sohrecs)'

    DeleteMatlsSelectSQL = "INSERT INTO WICS_tmpmateriallistupdate (recStatus, delMaterialLink, MaterialLink_id, org_id, Material, Description, Plant "
    DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  ) "    # these can go once I set null=True on these fields
    DeleteMatlsSelectSQL += " SELECT  concat('DEL ',FORMAT(id,0)), id, NULL, org_id, Material, Description, Plant "
    DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  "    # these can go once I set null=True on these fields
    DeleteMatlsSelectSQL += " FROM WICS_materiallist"
    DeleteMatlsSelectSQL += f" WHERE ({MustKeepMatlsSelCond})"
    with connections[dbToUse].cursor() as cursor:
        cursor.execute(DeleteMatlsSelectSQL)

    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'id-add-matl',
        statetext = f'Identifying SAP MM60 Materials new to WICS',
        )
    MarkAddMatlsSelectSQL = "UPDATE WICS_tmpmateriallistupdate"
    MarkAddMatlsSelectSQL += " SET recStatus = 'ADD'"
    MarkAddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus is NULL)"
    with connections[dbToUse].cursor() as cursor:
        cursor.execute(MarkAddMatlsSelectSQL)
        transaction.on_commit(partial(done_MatlListSAPSprsheet_02_identifyexistingMaterial,dbToUse,reqid))
def done_MatlListSAPSprsheet_02_identifyexistingMaterial(dbToUse, reqid):
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'get-matl-link-done',
        statetext = f'Finished linking SAP MM60 list to existing WICS Materials',
        )
    
    proc_MatlListSAPSprsheet_03_UpdateExistingRecs(dbToUse, reqid)

def proc_MatlListSAPSprsheet_03_UpdateExistingRecs(dbToUse, reqid):
    def setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(dbToUse, fldName):
        acomm = set_async_comm_state(
            dbToUse,
            reqid,
            statecode = 'upd-existing-recs',
            statetext = f'Updating _{fldName}_ Field in Existing Records',
            )

    setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(dbToUse, '')

    # (Form Name, db fld Name, zero/blank value)
    FormTodbFld_map = [
        ('Description','Description','""'),
        ('SAPMatlType','SAPMaterialType','""'),
        ('SAPMatlGroup','SAPMaterialGroup','""'),
        ('SAPManuf','SAPManuf','""'),
        ('SAPMPN','SAPMPN','""'),
        ('SAPABC','SAPABC','""'),
        ('SAPPrice','Price',0),
        ('SAPPrice','PriceUnit',0),
        ('SAPPrice','Currency','""'),
    ]

    UpdateExistFldList_str = async_comm.objects.using(dbToUse).get(pk=f"{reqid}-UpdExstFldList").statetext
    UpdateExistFldList = ast.literal_eval(UpdateExistFldList_str)

    if UpdateExistFldList:
        for formName, dbName, zeroVal in FormTodbFld_map:
            if formName in UpdateExistFldList:
                setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(dbToUse, dbName)
                # UPDATE this field
                UpdSQLSetStmt = f"MatlList.{dbName}=tmpMatl.{dbName}"
                UpdSQLWhereStmt = f"(IFNULL(tmpMatl.{dbName},{zeroVal}) != {zeroVal} AND IFNULL(MatlList.{dbName},{zeroVal})!=IFNULL(tmpMatl.{dbName},{zeroVal}))"

                UpdSQLStmt = "UPDATE WICS_materiallist AS MatlList, WICS_tmpmateriallistupdate AS tmpMatl"
                UpdSQLStmt += f" SET {UpdSQLSetStmt}"
                UpdSQLStmt += f" WHERE (tmpMatl.MaterialLink_id=MatlList.id) AND {UpdSQLWhereStmt}"
                with connections[dbToUse].cursor() as cursor:
                    cursor.execute(UpdSQLStmt)
            #endif formName in UpdateExistFldList
        #endfor
    # endif UpdateExistFldList not empty
    done_MatlListSAPSprsheet_03_UpdateExistingRecs(dbToUse, reqid)
def done_MatlListSAPSprsheet_03_UpdateExistingRecs(dbToUse, reqid):
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'upd-existing-recs-done',
        statetext = f'Finished Updating Existing Records to MM60 values',
        )
    proc_MatlListSAPSprsheet_04_Remove(dbToUse, reqid)

def proc_MatlListSAPSprsheet_04_Remove(dbToUse, reqid):
    ## MustKeepMatlsDelCond = ''
    ## if MustKeepMatlsDelCond: MustKeepMatlsDelCond += ' AND '
    ## MustKeepMatlsDelCond += 'id IN (SELECT DISTINCT delMaterialLink FROM WICS_tmpmateriallistupdate WHERE recStatus like "DEL%")'

    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'del-matl',
        statetext = f'Removing WICS Materials no longer in SAP MM60 Materials',
        )
    # do the Removals
    ## DeleteMatlsDoitSQL = "DELETE FROM WICS_materiallist"
    ## DeleteMatlsDoitSQL += f" WHERE ({MustKeepMatlsDelCond})"
    DeleteMatlsDoitSQL = 'DELETE MATL'
    DeleteMatlsDoitSQL += ' FROM WICS_materiallist AS MATL INNER JOIN WICS_tmpmateriallistupdate AS TMP'
    DeleteMatlsDoitSQL += '    ON MATL.id = TMP.delMaterialLink'
    DeleteMatlsDoitSQL += ' WHERE TMP.recStatus like "DEL%"'
    with connections[dbToUse].cursor() as cursor:
        cursor.execute(DeleteMatlsDoitSQL)
        transaction.on_commit(partial(done_MatlListSAPSprsheet_04_Remove,dbToUse,reqid))
def done_MatlListSAPSprsheet_04_Remove(dbToUse,reqid):
    mandatorytaskdonekey = f'MatlX{reqid}'
    statecodeVal = ".03D."
    existingstatecode = ''
    if async_comm.objects.using(dbToUse).filter(pk=mandatorytaskdonekey).exists(): existingstatecode = async_comm.objects.using(dbToUse).get(pk=mandatorytaskdonekey).statecode
    MatlXval = existingstatecode + statecodeVal
    set_async_comm_state(
        dbToUse,
        mandatorytaskdonekey,
        statecode = MatlXval,
        statetext = '',
        new_async = True
        )
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'done-del-matl',
        statetext = f'Finished Removing WICS Materials no longer in SAP MM60 Materials',
        )
    proc_MatlListSAPSprsheet_04_Add(dbToUse, reqid)

def proc_MatlListSAPSprsheet_04_Add(dbToUse, reqid):
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'add-matl',
        statetext = f'Adding SAP MM60 Materials new to WICS',
        )
    UnknownTypeID = WhsePartTypes.objects.using(dbToUse).get(WhsePartType=WICS.globals._PartTypeName_UNKNOWN)
    # do the adds
    # one day django will implement insert ... select.  Until then ...
    AddMatlsSelectSQL = "SELECT"
    AddMatlsSelectSQL += " org_id, Material, Description, Plant, " + str(UnknownTypeID.pk) + " AS PartType_id,"
    AddMatlsSelectSQL += " SAPMaterialType, SAPMaterialGroup, Price, PriceUnit, Currency,"
    AddMatlsSelectSQL += " '' AS TypicalContainerQty, '' AS TypicalPalletQty, '' AS Notes"
    AddMatlsSelectSQL += " FROM WICS_tmpmateriallistupdate"
    AddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus = 'ADD') "

    AddMatlsDoitSQL = "INSERT INTO WICS_materiallist"
    AddMatlsDoitSQL += " (org_id, Material, Description, Plant, PartType_id,"
    AddMatlsDoitSQL += " SAPMaterialType, SAPMaterialGroup, Price, PriceUnit, Currency,"
    AddMatlsDoitSQL += " TypicalContainerQty, TypicalPalletQty, Notes)"
    AddMatlsDoitSQL += " " + AddMatlsSelectSQL
    with connections[dbToUse].cursor() as cursor:
        cursor.execute(AddMatlsDoitSQL)

    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'add-matl-get-recid',
        statetext = f'Getting Record ids of SAP MM60 Materials new to WICS',
        )
    UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
    UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id '
    UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
    UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
    UpdMaterialLinkSQL += "   and (MaterialLink_id IS NULL) AND (recStatus = 'ADD')"
    with connections[dbToUse].cursor() as cursor:
        cursor.execute(UpdMaterialLinkSQL)
        transaction.on_commit(partial(done_MatlListSAPSprsheet_04_Add,dbToUse,reqid))
def done_MatlListSAPSprsheet_04_Add(dbToUse, reqid):
    mandatorytaskdonekey = f'MatlX{reqid}'
    statecodeVal = ".03A."
    existingstatecode = ''
    if async_comm.objects.using(dbToUse).filter(pk=mandatorytaskdonekey).exists(): existingstatecode = async_comm.objects.using(dbToUse).get(pk=mandatorytaskdonekey).statecode
    MatlXval = existingstatecode + statecodeVal
    set_async_comm_state(
        dbToUse,
        mandatorytaskdonekey,
        statecode = MatlXval,
        statetext = '',
        new_async = True
        )
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'done-add-matl',
        statetext = f'Finished Adding SAP MM60 Materials new to WICS',
        )

def proc_MatlListSAPSprsheet_99_FinalProc(dbToUse, reqid):
    set_async_comm_state(
        dbToUse,
        reqid,
        statecode = 'done',
        statetext = 'Finished Processing Spreadsheet',
        )
    
def proc_MatlListSAPSprsheet_99_Cleanup(dbToUse, reqid):
    # also kill reqid, acomm, qcluster process
    async_comm.objects.using(dbToUse).filter(pk=reqid).delete()
    async_comm.objects.using(dbToUse).filter(pk=f"{reqid}-UpdExstFldList").delete()

    # when we can start django-q programmatically, this is where we kill that process
    try:
        os.kill(int(reqid), signal.SIGTERM)
    except AttributeError:
        pass
    try:
        os.kill(int(reqid), signal.SIGKILL)
    except AttributeError:
        pass

    # delete the temporary table
    tmpMaterialListUpdate.objects.using(dbToUse).all().delete()

@login_required
def fnUpdateMatlListfromSAP(req):

    dbToUse = user_db(req)

    client_phase = req.POST['phase'] if 'phase' in req.POST else None
    reqid = req.COOKIES['reqid'] if 'reqid' in req.COOKIES else None

    if req.method == 'POST':
        # check if the mandatory commits have been done and change the status code if so
        if reqid is not None:
            mandatory_commit_key = f'MatlX{reqid}'
            mandatory_commit_list = ['03A', '03D']
            if async_comm.objects.using(dbToUse).filter(pk=mandatory_commit_key).exists():
                mandatory_commits_recorded = async_comm.objects.using(dbToUse).get(pk=mandatory_commit_key).statecode
                if all((c in mandatory_commits_recorded) for c in mandatory_commit_list):
                    proc_MatlListSAPSprsheet_99_FinalProc(dbToUse, reqid)
                    async_comm.objects.using(dbToUse).filter(pk=mandatory_commit_key).delete()

        retinfo = HttpResponse()
        if client_phase=='init-upl':
            # start django_q broker
            reqid = subprocess.Popen(
                ['python', f'{django_settings.BASE_DIR}/manage.py', 'qcluster']
            ).pid
            # reqid = random.randint(1, 100000000000)
            retinfo.set_cookie('reqid',str(reqid))

            UpdateExistFldList = req.POST.getlist('UpIfCh', default=[])
            proc_MatlListSAPSprsheet_00InitUMLasync_comm(dbToUse, reqid, UpdateExistFldList)

            UMLSSName = proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(req, reqid)
            async_task(proc_MatlListSAPSprsheet_01ReadSpreadsheet, dbToUse, reqid, UMLSSName, hook=done_MatlListSAPSprsheet_01ReadSpreadsheet)

            acomm = async_comm.objects.using(dbToUse).values().get(pk=reqid)    # something's very wrong if this doesn't exist
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='waiting':
            acomm = async_comm.objects.using(dbToUse).values().get(pk=reqid)    # something's very wrong if this doesn't exist
            retinfo.write(json.dumps(acomm))
            return retinfo
        elif client_phase=='wantresults':
            ImpErrList = tmpMaterialListUpdate.objects.using(dbToUse).filter(recStatus__startswith='err')
            AddedMatlsList = tmpMaterialListUpdate.objects.using(dbToUse).filter(recStatus='ADD')
            RemvdMatlsList = tmpMaterialListUpdate.objects.using(dbToUse).filter(recStatus__startswith='DEL')
            cntext = {
                'ImpErrList':ImpErrList,
                'AddedMatls':AddedMatlsList,
                'RemvdMatls':RemvdMatlsList,
                }
            templt = 'frmUpdateMatlListfromSAP_done.html'
            return render(req, templt, cntext)
        elif client_phase=='cleanup-after-failure':
            pass
        elif client_phase=='resultspresented':
            proc_MatlListSAPSprsheet_99_Cleanup(dbToUse, reqid)
            retinfo.delete_cookie('reqid')

            return retinfo
        else:
            return
        #endif client_phase
    else:   # req.method != 'POST'
        # (hopefully,) this is the initial phase; all others will be part of a POST request

        cntext = {
            'reqid': -1,
            }
        templt = 'frmUpdateMatlListfromSAP_phase0.html'
    #endif req.method = 'POST'

    return render(req, templt, cntext)
