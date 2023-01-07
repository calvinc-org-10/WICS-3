from userprofiles.models import WICSuser
from WICS.models import WhsePartTypes, MaterialList, tmpMaterialListUpdate
from cMenu.models import getcParm
from django.shortcuts import render
from django.core.files.uploadedfile import TemporaryUploadedFile, InMemoryUploadedFile
from openpyxl import load_workbook
from io import open
import os
import uuid

"""
    qs = SAP
    org = _userorg
    Material = Material
    Description = Material description
    PartType = get(UNKNOWN)
    SAPMaterialType = Material type
    SAPMaterialGroup = Material Group
    Price = Price
    PriceUnit = Price unit
"""

def fnUpdateMatlListfromSAP(req):
    _userorg = WICSuser.objects.get(user=req.user).org

    if req.method == 'POST':
        if req.POST['NextPhase']=='02-Upl-Sprsht':
            SAPFile = req.FILES['SAPFile']
            svdir = getcParm('SAP-FILELOC')
            fName = svdir+"tmpMatlList"+str(uuid.uuid4())+".xlsx"
            with open(fName, "wb") as destination:
                for chunk in SAPFile.chunks():
                    destination.write(chunk)

            tmpMaterialListUpdate.objects.all().delete()

            wb = load_workbook(filename=fName, read_only=True)
            ws = wb.active
            SAPcolmnNames = ws[1]
            SAPcol = {}
            SAPcolmap = {
                    'Material': 'Material', 
                    'Material description': 'Description', 
                    'Material type': 'SAPMaterialType',
                    'Material Group': 'SAPMaterialGroup',
                    'Price': 'Price',
                    'Price unit': 'PriceUnit',
                    }
            for col in SAPcolmnNames:
                if col.value in SAPcolmap:
                    SAPcol[SAPcolmap[col.value]] = col.column - 1
            #if (SAPcol['Material'] == None or SAPcol['StorageLocation'] == None or SAPcol['Amount'] == None):
            #    raise Exception('SAP Spreadsheet has bad header row.  See Calvin to fix this.')

            for row in ws.iter_rows(min_row=2, values_only=True):
                tmpMaterialListUpdate(
                                Material = row[SAPcol['Material']], 
                                Description = row[SAPcol['Description']], 
                                SAPMaterialType = row[SAPcol['SAPMaterialType']],
                                SAPMaterialGroup = row[SAPcol['SAPMaterialGroup']],
                                Price = row[SAPcol['Price']],
                                PriceUnit = row[SAPcol['PriceUnit']]
                                ).save()
            # endfor

            
            # later, save (FileMatList - MaterialList) and (MaterialList - FileMatList)
            # ask permission to correct each to the other
            # that will involve a temp table 
            # almost there now that I've created tmpMaterialListUpdate

            # for now ...
            AddedMatls = tmpMaterialListUpdate.objects.exclude(Material__in=MaterialList.objects.filter(org=_userorg).values('Material'))
            # one day django will implement insert ... select.  Until then ...
            for newRec in AddedMatls:
                MaterialList (
                        org = _userorg,
                        Material = newRec.Material,
                        Description = newRec.Description,
                        PartType = WhsePartTypes.objects.get(WhsePartType='UNKNOWN'),
                        SAPMaterialType = newRec.SAPMaterialType,
                        SAPMaterialGroup = newRec.SAPMaterialGroup,
                        Price = newRec.Price,
                        PriceUnit = newRec.PriceUnit
                        ).save()

            # delete the temporary table and the temporary file
            tmpMaterialListUpdate.objects.all().delete()
            wb.close()
            os.remove(fName)

        # endif req.POST['NextPhase']=='02-Upl-Sprsht'
        cntext = {'AddedMatls':AddedMatls}
        templt = 'frmUpdateMatlListfromSAP_done.html'
    else:
        # (hopefully,) this is the initial phase; all others will be part of a POST request
        cntext = {}
        templt = 'frmUpdateMatlListfromSAP_phase0.html'
    #endif req.method = 'POST'

    cntext['orgname'] = _userorg.orgname
    cntext['uname'] = req.user.get_full_name()
    return render(req, templt, cntext)

