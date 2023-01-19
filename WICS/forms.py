### TODO: Split into procs_Material, procs_ActualCounts, procs_ScheduledCounts, procs_SAPData
###       fold in reports.py, too

import datetime
import os, uuid
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.urls import reverse
from django import forms
from django.db import models
from django.forms import inlineformset_factory, formset_factory
from django.shortcuts import render
from django.db.models import Value, Sum
from cMenu.models import getcParm
from WICS.models import MaterialList, ActualCounts, CountSchedule, \
                        SAP_SOHRecs, \
                        WhsePartTypes, Organizations, LastFoundAt
from WICS.SAPLists import fnSAPList
from userprofiles.models import WICSuser
from django.http import HttpResponseRedirect, HttpResponse, HttpRequest
from django.db.models.query import EmptyQuerySet, QuerySet
from django.views.generic import ListView
from openpyxl import load_workbook
from typing import Any, Dict


ExcelWorkbook_fileext = ".XLSX"
#_userorg = WICSuser.objects.none().org
_userorg = None


class MaterialForm(forms.ModelForm):
    showPK = forms.IntegerField(label='ID', disabled=True, required=False)
    class Meta:
        model = MaterialList
        fields = ['id', 'org', 'Material', 'Description','PartType',
                'SAPMaterialType', 'SAPMaterialGroup', 'Price',
                'TypicalContainerQty', 'TypicalPalletQty', 'PriceUnit', 'Notes']
        # fields = '__all__'

class MaterialCountSummary(forms.Form):
    Material = forms.CharField(max_length=100, disabled=True)
    CountDate = forms.DateField(required=False, disabled=True)
    CountQTY_Eval = forms.IntegerField(required=False, disabled=True)


@login_required
def fnMaterialForm(req, recNum = -1, gotoRec=False):
    _userorg = WICSuser.objects.get(user=req.user).org
    if not _userorg: raise Exception('User is corrupted!!')

    # get current record
    currRec = None
    if gotoRec:
        currRec = MaterialList.objects.filter(org=_userorg, Material=req.GET['gotoID']).first()
    if not currRec:
        if recNum <= 0:
            currRec = MaterialList.objects.filter(org=_userorg).first()
        else:
            currRec = MaterialList.objects.filter(org=_userorg).get(pk=recNum)   # later, handle record not found
        # endif
    #endif
    if not currRec: #there are no MaterialList records for this org!!
        thisPK = 0
    else:
        thisPK = currRec.pk

    SAP_SOH = fnSAPList(_userorg, matl=currRec.Material)

    gotoForm = {}
    gotoForm['gotoItem'] = currRec
    gotoForm['choicelist'] = MaterialList.objects.filter(org=_userorg).values('id','Material')

    changes_saved = {
        'main': False,
        'counts': False,
        'schedule': False
        }
    chgd_dat = {'main':None, 'counts': None, 'schedule': None}

    CountSubFormFields = ('id', 'CountDate', 'CycCtID', 'Counter', 'LocationOnly', 'CTD_QTY_Expr', 'BLDG', 'LOCATION', 'PKGID_Desc', 'TAGQTY', 'FLAG_PossiblyNotRecieved', 'FLAG_MovementDuringCount', 'Notes',)
    ScheduleSubFormFields = ('id','CountDate','Counter', 'Priority', 'ReasonScheduled', 'CMPrintFlag', 'Notes',)

    if req.method == 'POST':
        # changed data is being submitted.  process and save it
        # process mtlFm AND subforms.

        # process main form
        #if currRec:
        mtlFm = MaterialForm(req.POST, instance=currRec,  initial={'gotoItem': thisPK, 'showPK': thisPK, 'org':_userorg},  prefix='material')
        mtlFm.fields['PartType'].queryset=WhsePartTypes.objects.filter(org=_userorg).order_by('WhsePartType').all()

        #else:
        #    mtlFm = MaterialForm(req.POST, initial={'gotoItem': thisPK, 'showPK': thisPK, 'org':_userorg},  prefix='material')
        #endif
        if mtlFm.is_valid():
            if mtlFm.has_changed():
                mtlFm.save()
                chgd_dat['main'] = mtlFm.changed_data
                changes_saved['main'] = True
                #raise Exception('main saved')

        # count detail subform
        countSubFm_class = inlineformset_factory(MaterialList,ActualCounts,
                    fields=CountSubFormFields,
                    extra=0,can_delete=False)
        #if currRec:
        countSet = countSubFm_class(req.POST, instance=currRec, prefix='countset', initial={'org': _userorg}, queryset=ActualCounts.objects.order_by('-CountDate'))
        #else:
        #    countSet = countSubFm_class(req.POST, prefix='countset', initial={'org': _userorg}, queryset=ActualCounts.objects.order_by('-CountDate'))
        if countSet.is_valid():
            if countSet.has_changed():
                countSet.save()
                chgd_dat['counts'] = countSet.changed_objects
                changes_saved['counts'] = True
                #raise Exception('counts saved')

        # count schedule subform
        SchedSubFm_class = inlineformset_factory(MaterialList,CountSchedule,
                    fields=ScheduleSubFormFields,
                    extra=0,can_delete=False)
        #if currRec:
        schedSet = SchedSubFm_class(req.POST, instance=currRec, prefix='schedset', initial={'org': _userorg}, queryset=CountSchedule.objects.order_by('-CountDate'))
        #else:
        #    schedSet = SchedSubFm_class(req.POST, prefix='schedset', initial={'org': _userorg}, queryset=CountSchedule.objects.order_by('-CountDate'))
        if schedSet.is_valid():
            if schedSet.has_changed():
                schedSet.save()
                chgd_dat['schedule'] = schedSet.changed_objects
                changes_saved['schedule'] = True
                #raise Exception('sched saved')

        # count summary form is r/o.  It will not be changed
    else: # request.method == 'GET' or something else
        #if currRec:
        mtlFm = MaterialForm(instance=currRec, initial={'gotoItem': thisPK, 'showPK': thisPK, 'org':_userorg}, prefix='material')
        mtlFm.fields['PartType'].queryset=WhsePartTypes.objects.filter(org=_userorg).order_by('WhsePartType').all()
        #else:
        #    mtlFm = MaterialForm(initial={'gotoItem': thisPK, 'showPK': thisPK, 'org':_userorg}, prefix='material')

        CountSubFm_class = inlineformset_factory(MaterialList,ActualCounts,
                    fields=CountSubFormFields,
                    extra=0,can_delete=False)
        #if currRec:
        countSet = CountSubFm_class(instance=currRec, prefix='countset', initial={'org':_userorg}, queryset=ActualCounts.objects.order_by('-CountDate'))
        #else:
        #    countSet = CountSubFm_class(prefix='countset', initial={'org':_userorg}, queryset=ActualCounts.objects.order_by('-CountDate'))

        SchedSubFm_class = inlineformset_factory(MaterialList,CountSchedule,
                    fields=ScheduleSubFormFields,
                    extra=0,can_delete=False)
        #if currRec:
        schedSet = SchedSubFm_class(instance=currRec, prefix='schedset', initial={'org':_userorg}, queryset=CountSchedule.objects.order_by('-CountDate'))
        #else:
        #    schedSet = SchedSubFm_class(prefix='schedset', initial={'org':_userorg}, queryset=CountSchedule.objects.order_by('-CountDate'))
    # endif

    # count summary subform
    raw_countdata = ActualCounts.objects.filter(Material=currRec).order_by('Material','-CountDate').annotate(QtyEval=Value(0, output_field=models.IntegerField()))
    LastMaterial = None ; LastCountDate = None
    initdata = []
    for r in raw_countdata:
        try:
            r.QtyEval = eval(r.CTD_QTY_Expr)    # later, use ast.literal_eval
        # except (ValueError, SyntaxError):
        except:
            r.QtyEval = 0
        if (r.Material != LastMaterial or r.CountDate != LastCountDate):
            LastMaterial = r.Material ; LastCountDate = r.CountDate
            initdata.append({
                'Material': r.Material,
                'CountDate': r.CountDate,
                'CountQTY_Eval': 0,
            })
        n = initdata[-1]['CountQTY_Eval'] + r.QtyEval
        initdata[-1]['CountQTY_Eval'] = n
    subFm_class = formset_factory(MaterialCountSummary,extra=0)
    summarySet = subFm_class(initial=initdata, prefix='summaryset')

    #countSet['org'].is_hidden = True
    #schedSet['org'].is_hidden = True

    # display the form
    cntext = {'frmMain': mtlFm,
            'lastFoundAt': LastFoundAt(currRec),
            'gotoForm': gotoForm,
            'countset': countSet,
            'scheduleset': schedSet,
            'countsummset': summarySet,
            'SAPSet': SAP_SOH,
            'changes_saved': changes_saved,
            'changed_data': chgd_dat,
            'recNum': recNum,
            'orgname':_userorg.orgname, 'uname':req.user.get_full_name()
            }
    templt = 'frm_Material.html'
    return render(req, templt, cntext)

#####################################################################################################
#####################################################################################################
#####################################################################################################

class UploadSAPForm(forms.Form):
    uploaded_at = forms.DateTimeField()
    SAPFile = forms.FileField()

@login_required
def fnUploadSAP(req):
    _userorg = WICSuser.objects.get(user=req.user).org

    if req.method == 'POST':
        form = UploadSAPForm(req.POST, req.FILES)
        if form.is_valid():
            # save the file so we can open it as an excel file
            SAPFile = req.FILES['SAPFile']
            svdir = getcParm('SAP-FILELOC')
            fName = svdir+"tmpSAP"+str(uuid.uuid4())+".xlsx"
            with open(fName, "wb") as destination:
                for chunk in SAPFile.chunks():
                    destination.write(chunk)

            wb = load_workbook(filename=fName, read_only=True)
            ws = wb.active

            # I map Table Fields directly to spreadsheet columns because the SOH spreadsheets
            # are fairly stable.  If that changes, see fnUpdateMatlListfromSAP in SAPMatlUpdt.py
            # for an alternative way of handling this mapping
            SAPcolmnMap = {
                        'Material': 0,
                        'Description': 1,
                        'Plant': 2,
                        'MaterialType': 3,
                        'StorageLocation': 4,
                        'BaseUnitofMeasure': 5,
                        'Amount': 6,
                        'Currency': 7,
                        'ValueUnrestricted': 8,
                        'SpecialStock': 9,
                        'Batch': 10,
                        }

            nRows = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[SAPcolmnMap['Material']]==None: MatNum = ''
                else: MatNum = row[SAPcolmnMap['Material']]
                if len(str(MatNum)):
                    SRec = SAP_SOHRecs(
                                org = _userorg,
                                uploaded_at = req.POST['uploaded_at']
                                )
                    for fldName, colNum in SAPcolmnMap.items():
                        if row[colNum]==None: setval = ''
                        else: setval = row[colNum]
                        setattr(SRec, fldName, setval)
                    SRec.save()
                    nRows += 1

            # close and kill temp files
            wb.close()
            os.remove(fName)

            cntext = {'uploaded_at':req.POST['uploaded_at'], 'nRows':nRows,
                    'orgname':_userorg.orgname, 'uname':req.user.get_full_name()
                    }
            templt = 'frm_upload_SAP_Success.html'
    else:
        form = UploadSAPForm()
        cntext = {'form': form, 
                'orgname':_userorg.orgname, 'uname':req.user.get_full_name()
                }
        templt = 'frm_upload_SAP.html'
    #endif

    return render(req, templt, cntext)

#####################################################################################################
#####################################################################################################
#####################################################################################################



#####################################################################
#####################################################################
#####################################################################

class CountScheduleForm(ListView):
    ordering = ['-CountDate', 'Material']
    context_object_name = 'CtSchdList'
    template_name = 'frm_CountScheduleList.html'
    
    def setup(self, req: HttpRequest, *args: Any, **kwargs: Any) -> None:
        self._user = req.user
        self._userorg = WICSuser.objects.get(user=req.user).org
        self.queryset = CountSchedule.objects.filter(org=self._userorg).order_by('-CountDate', 'Material')   # figure out how to pass in self.ordering
        return super().setup(req, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Any]:
        return super().get_queryset()

    # def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return super().get(request, *args, **kwargs)

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        return HttpResponse('Stop rushing me!!')

    def render_to_response(self, context: Dict[str, Any], **response_kwargs: Any) -> HttpResponse:
        context.update({'orgname': self._userorg.orgname,  'uname':self._user.get_full_name()})
        return super().render_to_response(context, **response_kwargs)


class MaterialByPartType(ListView):
    ordering = ['PartType__PartTypePriority', 'Material']
    context_object_name = 'MatlList'
    template_name = 'frm_MatlByPartTypeList.html'
    SAPSums = {}
    
    def setup(self, req: HttpRequest, *args: Any, **kwargs: Any) -> None:
        self._user = req.user
        self._userorg = WICSuser.objects.get(user=self._user).org
        self.queryset = MaterialList.objects.filter(org=self._userorg).order_by('PartType__PartTypePriority', 'Material').annotate(LFADate=Value(0), LFALocation=Value(''), enumerate_in_group=Value(0), SAPQty=Value(0))   # figure out how to pass in self.ordering
        
        # it's more efficient to pull this all now and store it for the upcoming qs request
        SAP = fnSAPList(self._userorg)
        self.SAPDate = SAP['SAPDate']
        rawsums = SAP['SAPTable'].values('Material').annotate(TotalAmount=Sum('Amount',default=0))
        for x in rawsums: self.SAPSums[x['Material']] = x['TotalAmount']
        
        return super().setup(req, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Any]:
        qs = super().get_queryset()
        LastPT = None
        enumInGrp = 0
        for rec in qs:
            enumInGrp += 1
            L = LastFoundAt(rec)
            rec.LFADate = L['lastCountDate']
            rec.LFALocation = L['lastFoundAt']
            if rec.PartType != LastPT:
                enumInGrp = 1
                LastPT = rec.PartType
            rec.enumerate_in_group = enumInGrp
            rec.SAPQty = 0
            if rec.Material in self.SAPSums: rec.SAPQty = self.SAPSums[rec.Material]

        return qs

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        ctxt = super().get_context_data(**kwargs)
        ctxt['SAPDate'] = self.SAPDate
        return ctxt

    # def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return super().get(request, *args, **kwargs)

    def render_to_response(self, context: Dict[str, Any], **response_kwargs: Any) -> HttpResponse:
        context.update({'orgname': self._userorg.orgname,  'uname':self._user.get_full_name()})
        return super().render_to_response(context, **response_kwargs)


