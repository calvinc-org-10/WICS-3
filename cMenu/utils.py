from __future__ import annotations      # to allow calvindate (or any other class) to refer to itself.  See https://stackoverflow.com/questions/40925470/python-how-to-refer-a-class-inside-itself
import datetime
from datetime import date, timedelta
from dateutil.parser import parse
from dateutil.rrule import *
from django.contrib.auth.models import User, Group
from django.shortcuts import render
from django.http import HttpRequest
from django.http.response import HttpResponse, HttpResponseForbidden
from django.db.models import QuerySet
from django.db.models import Aggregate, CharField
from collections import namedtuple
from cMenu.externalWebPageURL_Map import ExternalWebPageURL_Map
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, fills, colors
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH


ExcelWorkbook_fileext = ".XLSX"


class calvindate(date):
    def __new__(cls, *args):
        D = datetime.date.today()        # set default to datetime.date.today()
        if len(args) == 3:  # year, month, day was passed in
            return super().__new__(cls, int(args[0]), int(args[1]), int(args[2]))
        elif len(args) == 2:    # month, day passed in , year should be current year
            yrr = datetime.date.today().year
            return super().__new__(cls, int(yrr), int(args[0]), int(args[1]))
        elif len(args) == 1:    # either a date string or date object passed in
            if isinstance(args[0],(datetime.date, datetime.datetime)):
                DStr = str(args[0])
            else:
                DStr = args[0]
            try:
                D = parse(DStr)
                D = D.date()
            except:
                pass
        else:
            # invalid number of args.  Do nothing; let the default stand
            pass
        return super().__new__(cls, D.year, D.month, D.day)

    def as_datetime(self):
        return datetime.datetime(self.year,self.month,self.day)

    def today(self):
        return self.__class__()
    def daysfrom(self,delta:int) -> calvindate:
        R_dt = self.as_datetime() + timedelta(days=delta)
        return calvindate(R_dt)
    def tomorrow(self):
        return self.daysfrom(1)
    def yesterday(self):
        return self.daysfrom(-1)
    
    def nextWorkdayAfter(self, nonWorkdays={SA,SU}, extraNonWorkdayList={}, include_afterdate=False):
        afterdate = self.as_datetime()
        
        excRule = rrule(WEEKLY,dtstart=afterdate,byweekday=nonWorkdays)
        afterdaysRule = rrule(DAILY,dtstart=afterdate)

        exclSet = rruleset()
        exclSet.rrule(afterdaysRule)
        exclSet.exrule(excRule)
        # loop extraNonWorkdays into exclSet.exdate
        for xDate in extraNonWorkdayList:
            exclSet.exdate(xDate)

        return self.__class__( exclSet.after(afterdate,include_afterdate) )
    
    # operators
    def __comparison_workhorse__(self, RHE, compOpr):
        LHExpr = calvindate(self).as_datetime()
        RHExpr = calvindate(RHE).as_datetime()
        if compOpr == 'lt':
            return LHExpr < RHExpr
        if compOpr == 'le':
            return LHExpr <= RHExpr
        if compOpr == 'eq':
            return LHExpr == RHExpr
        if compOpr == 'ne':
            return LHExpr != RHExpr
        if compOpr == 'gt':
            return LHExpr > RHExpr
        if compOpr == 'ge':
            return LHExpr >= RHExpr
        return False
    def __lt__(self, other):
        return self.__comparison_workhorse__(other,'lt')
    def __le__(self, other):
        return self.__comparison_workhorse__(other,'le')
    def __eq__(self, other):
        return self.__comparison_workhorse__(other,'eq')
    def __ne__(self, other):
        return self.__comparison_workhorse__(other,'ne')
    def __gt__(self, other):
        return self.__comparison_workhorse__(other,'gt')
    def __ge__(self, other):
        return self.__comparison_workhorse__(other,'ge')
    def __add__(self, other):
        if isinstance(other, int):
            return self.daysfrom(other)
        else:
            return NotImplemented
    def __sub__(self, other):
        if isinstance(other, int):
            return self.daysfrom(-other)
        else:
            return NotImplemented


def WrapInQuotes(strg, openquotechar = '"', closequotechar = '"'):
    return openquotechar + strg + closequotechar


def makebool(strngN, numtype = bool):
    # res = bool(strngN)
    # the built-in bool function is good with one set of exceptions
    if isinstance(strngN,str):
        FalsieList = ['FALSE','NO','0','OFF','']
        if  (strngN.strip().upper() in FalsieList) or len(strngN)<1:
            strngN = False
    else:
        strngN = numtype(strngN)
    return strngN
def iif(cond, ifTrue, ifFalse=None):
    if makebool(cond):
        return ifTrue
    else:
        return ifFalse


def isDate(testdate):
    try:
        td = parse(testdate)
    except:
        td = False
    return td


def dictfetchall(cursor):
    """
    Return all rows from a cursor as a list of dictionaries.
    Assume the column names are unique.
    """
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

def namedtuplefetchall(cursor, ResultName = 'Result'):
    "Return all rows from a cursor as a namedtuple"
    desc = cursor.description
    nt_result = namedtuple(ResultName, [col[0] for col in desc])
    return [nt_result(*row) for row in cursor.fetchall()]

def rgb_to_int(red,green,blue):
    rgb = red
    rgb = (rgb << 8) + green
    rgb = (rgb << 8) + blue
    return rgb
def int_to_rgb(rgb_int):
    red = (rgb_int >> 16) & 0xFF
    green = (rgb_int >> 8) & 0xFF
    blue = rgb_int & 0xFF
    return (red,green,blue)

def modelobj_to_dict(modelobj):
    # opts = modelobj._meta
    # data = {}
    # for f in chain(opts.concrete_fields, opts.private_fields):
    #     data[f.name] = f.value_from_object(modelobj)
    # for f in opts.many_to_many:
    #     data[f.name] = [i.id for i in f.value_from_object(modelobj)]

    data = {key:val for key, val in modelobj.__dict__.items() if key not in ['_state']}
    return data

def Excelfile_fromqs(qset, flName, freezecols = 0):
    """
    qset: a queryset or list of dictionaries
    flName: the name of the file to be built (WITHOUT extension!).  It's stored on the server.  If it's to be dl'd, the caller does that
    freezecols = 0: the number of columns to freeze to the left
    The top row contains the field names, is always frozen, is bold and is shaded grey

    Returns the name of the Workbook file (with extension).  Once I start building in errorchecking and exceptions, other returns may be possible
    """

    # far easier to process a list of dictionaries, so...
    if isinstance(qset,QuerySet):
        qlist = qset.values()
    elif isinstance(qset,list):
        qlist = qset
    else:
        return None
    if qlist:
        if not isinstance(qlist[0],dict):
            # review this later ...
            try:
                qlist = [n.__dict__ for n in qlist]
            except:
                qlist = []

    # create empty workbook with an empty worksheet
    wb = Workbook()
    ws = wb.active

    # header row is names of columns
    if qlist:
        fields = list(qlist[0])
        ws.append(fields)

        # append each row
        for row in qlist:
            ws.append(list(row.values()))

        # make header row bold, shade it grey, freeze it
        # ws.show_gridlines = True  #TODO: Find out how to do this
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type=fills.FILL_SOLID,
                            start_color=colors.Color("00808080"),
                            end_color=colors.Color("00808080")
                            )
        #TODO: convert row1 and cols:freezecols to an address (A=0, B=1, C=2 etc) for line below
        ws.freeze_panes ='A2'
        #TODO: if freezecols passed, freeze them, too

    # save the workbook
    wb.save(flName+ExcelWorkbook_fileext)

    # close the workbook
    wb.close()

    # and return success code to the caller
    return flName+ExcelWorkbook_fileext


# From https://stackoverflow.com/questions/10340684/group-concat-equivalent-in-django
class GroupConcat(Aggregate):
    function = 'GROUP_CONCAT'
    allow_distinct = True
    template = '%(function)s(%(distinct)s%(expressions)s%(ordering)s%(separator)s)'

    def __init__(self, expression, distinct=False, ordering=None, separator=None, **extra):
        super().__init__(
            expression,
            distinct='DISTINCT ' if distinct else '',
            ordering=' ORDER BY %s' % ordering if ordering is not None else '',
            separator=' SEPARATOR "%s"' % separator if separator is not None else '',
            output_field=CharField(),
            **extra
        )


class UpldSprdsheet():
    TargetModel = None
    SprdsheetDateEpoch = WINDOWS_EPOCH

    def SprdsheetFldDescriptor_creator(self, ModelFldName, AllowedTypes):
        """
        ModelFldName (str): the name of the field in the TargetModel
        AllowedTypes: list of tuples (type, cleanproc).  empty list if any string allowed
        """
        return  {
            # 'SprdsheetName': None,    # nope, this will be the index of SprdsheetFlds
            'ModelFldName': ModelFldName,
            'AllowedTypes': AllowedTypes,     
        }
    
    SprdsheetFlds = {}  # key will be the SprdsheetName, value is a SprdsheetFldDescriptor

    def cleanupfld(self, fld, val):
        usefld = False
        cleanval = None
        
        if fld not in self.SprdsheetFlds:
            # just feed the value back
            usefld = True
            cleanval = val
        elif not self.SprdsheetFlds[fld]['AllowedTypes']:
            usefld = (val is not None)
            if usefld: cleanval = str(val)
        else:
            for type, cleanproc in self.SprdsheetFlds[fld]['AllowedTypes']:
                if isinstance(val, type):
                    usefld = True
                    cleanval = cleanproc(val)
                    break
                #endif instance(val, type)
            #endfor type, cleanproc
        #endif fld not in self.SprdsheetFlds

    def process_spreadsheet(self, SprsheetName):
        pass


def fn_LoadExtWebPage(req, extpageURL):

    templt = "cUtilLoadExt.html"
    cntext = {
        'extpageURL': ExternalWebPageURL_Map[extpageURL],
        }
    theForm = render(req, templt, cntext)

    return theForm

def is_user_in_group(user: User, group_name: str) -> bool:
    return user.groups.filter(name=group_name).exists()

def is_superman(usr: User) -> bool:
    return usr.is_superuser or is_user_in_group(usr,'superman')

def user_db(req: HttpRequest|User|str) -> str:
    dbUsing = 'default'
    usr = None

    if   isinstance(req, HttpRequest):
        usr = req.user
    elif isinstance(req, User):
        usr = req
    elif isinstance(req, str):
        dbUsing = req
    #endif isinstance(req)

    if usr:
        if is_user_in_group(usr,'demo'): 
            dbUsing = 'demo'
        else:
            dbUsing = 'default'
        
    return dbUsing

def unauthorized_function(req: HttpRequest) -> HttpResponse:
    # templt = ''
    # cntext = {}

    return HttpResponseForbidden('Unauthorized') # render(req, templt, cntext, status=403)

